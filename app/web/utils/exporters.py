"""CSV/XLSX export utilities."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def to_csv(data: list[dict[str, Any]], columns: list[str]) -> str:
    """Convert data to CSV string.

    Args:
        data: List of dictionaries to export
        columns: Column names to include (in order)

    Returns:
        CSV string

    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def to_xlsx(data: list[dict[str, Any]], columns: list[str], sheet_name: str = "Data") -> bytes:
    """Convert data to XLSX bytes.

    Args:
        data: List of dictionaries to export
        columns: Column names to include (in order)
        sheet_name: Name for the Excel sheet

    Returns:
        XLSX file as bytes

    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row with bold font
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
