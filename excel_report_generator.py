#!/usr/bin/env python3
"""Генератор Excel таблиц для DDS (движение денежных средств) и P&L отчетов
Комбинирует данные из API с шаблонами себестоимости и расходов
"""

import asyncio
import json
import logging
import os
from datetime import datetime, timedelta
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from cost_template_generator import CostTemplateGenerator
from expenses import ExpenseManager
from real_data_reports import RealDataFinancialReports

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Генератор Excel отчетов с DDS и P&L таблицами"""

    def __init__(self):
        self.reports = RealDataFinancialReports()
        self.expense_manager = ExpenseManager()
        self.cost_generator = CostTemplateGenerator()

        # Создаем директории для отчетов
        os.makedirs("/root/sovani_bot/excel_reports", exist_ok=True)
        os.makedirs("/root/sovani_bot/cost_data", exist_ok=True)

    async def generate_dds_excel_report(self, date_from: str, date_to: str) -> str:
        """Генерация Excel таблицы ДДС (движение денежных средств)

        Args:
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)

        Returns:
            str: Путь к созданному Excel файлу

        """
        try:
            logger.info(f"Генерация DDS Excel отчета за период {date_from} - {date_to}")

            # Получаем финансовые данные
            pnl_data = await self.reports.calculate_real_pnl(date_from, date_to)

            # Создаем Excel файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"DDS_report_{date_from}_{date_to}_{timestamp}.xlsx"
            filepath = f"/root/sovani_bot/excel_reports/{filename}"

            wb = openpyxl.Workbook()

            # Удаляем дефолтный лист
            wb.remove(wb.active)

            # Создаем лист ДДС
            dds_sheet = wb.create_sheet("ДДС отчет")
            self._create_dds_sheet(dds_sheet, pnl_data, date_from, date_to)

            # Создаем лист детализации по платформам
            details_sheet = wb.create_sheet("Детализация")
            self._create_dds_details_sheet(details_sheet, pnl_data)

            # Сохраняем файл
            wb.save(filepath)

            logger.info(f"DDS Excel отчет создан: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Ошибка создания DDS Excel отчета: {e}")
            raise

    def _create_dds_sheet(self, sheet, pnl_data: dict[str, Any], date_from: str, date_to: str):
        """Создание основного листа ДДС"""
        # Заголовок
        sheet["A1"] = "ОТЧЕТ О ДВИЖЕНИИ ДЕНЕЖНЫХ СРЕДСТВ"
        sheet["A2"] = f"Период: {date_from} - {date_to}"
        sheet["A3"] = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        # Стили для заголовков
        title_font = Font(size=14, bold=True)
        header_font = Font(size=11, bold=True)
        currency_font = Font(size=10)

        sheet["A1"].font = title_font
        sheet["A2"].font = header_font
        sheet["A3"].font = currency_font

        # Основная таблица ДДС
        current_row = 6

        # ПОСТУПЛЕНИЯ
        sheet[f"A{current_row}"] = "ПОСТУПЛЕНИЯ"
        sheet[f"A{current_row}"].font = header_font
        current_row += 1

        # Выручка по платформам
        wb_revenue = pnl_data["wb"]["revenue"]
        ozon_revenue = pnl_data["ozon"]["revenue"]
        total_revenue = pnl_data["total"]["revenue"]

        sheet[f"A{current_row}"] = "Выручка от продаж"
        sheet[f"B{current_row}"] = total_revenue
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        sheet[f"A{current_row}"] = "  - Wildberries"
        sheet[f"B{current_row}"] = wb_revenue
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        sheet[f"A{current_row}"] = "  - Ozon"
        sheet[f"B{current_row}"] = ozon_revenue
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # К доплате (конечная сумма к перечислению)
        wb_final = pnl_data["wb"].get("final_revenue", wb_revenue)
        ozon_final = pnl_data["ozon"].get("final_revenue", ozon_revenue)
        total_final = wb_final + ozon_final

        sheet[f"A{current_row}"] = "К поступлению на счет"
        sheet[f"B{current_row}"] = total_final
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        sheet[f"B{current_row}"].font = header_font
        current_row += 2

        # СПИСАНИЯ
        sheet[f"A{current_row}"] = "СПИСАНИЯ"
        sheet[f"A{current_row}"].font = header_font
        current_row += 1

        # Комиссии маркетплейсов
        total_commission = pnl_data["total"]["commission"]
        sheet[f"A{current_row}"] = "Комиссии маркетплейсов"
        sheet[f"B{current_row}"] = -total_commission
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Себестоимость товаров
        total_cogs = pnl_data["total"]["cogs"]
        sheet[f"A{current_row}"] = "Себестоимость проданных товаров"
        sheet[f"B{current_row}"] = -total_cogs
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Рекламные расходы
        total_advertising = pnl_data["total"]["advertising"]
        sheet[f"A{current_row}"] = "Рекламные расходы"
        sheet[f"B{current_row}"] = -total_advertising
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Операционные расходы
        total_opex = pnl_data["total"]["opex"]
        sheet[f"A{current_row}"] = "Операционные расходы"
        sheet[f"B{current_row}"] = -total_opex
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Логистические расходы
        wb_logistics = pnl_data["wb"].get("logistics_costs", 0)
        ozon_logistics = pnl_data["ozon"].get("logistics_costs", 0)
        total_logistics = wb_logistics + ozon_logistics

        sheet[f"A{current_row}"] = "Логистика и доставка"
        sheet[f"B{current_row}"] = -total_logistics
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Возвраты
        wb_returns = pnl_data["wb"].get("returns_count", 0) * 100  # Оценочная стоимость
        ozon_returns = pnl_data["ozon"].get("returns_cost", 0)
        total_returns = wb_returns + ozon_returns

        sheet[f"A{current_row}"] = "Возвраты и брак"
        sheet[f"B{current_row}"] = -total_returns
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 2

        # ИТОГО СПИСАНИЯ
        total_outflow = (
            total_commission
            + total_cogs
            + total_advertising
            + total_opex
            + total_logistics
            + total_returns
        )
        sheet[f"A{current_row}"] = "ИТОГО СПИСАНИЯ"
        sheet[f"B{current_row}"] = -total_outflow
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        sheet[f"A{current_row}"].font = header_font
        sheet[f"B{current_row}"].font = header_font
        current_row += 2

        # ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК
        net_cashflow = total_final - total_outflow
        sheet[f"A{current_row}"] = "ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК"
        sheet[f"B{current_row}"] = net_cashflow
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"

        # Стилизация итоговой строки
        sheet[f"A{current_row}"].font = Font(size=12, bold=True)
        sheet[f"B{current_row}"].font = Font(size=12, bold=True)

        # Цвет фона в зависимости от результата
        if net_cashflow > 0:
            fill = PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )  # Зеленый
        else:
            fill = PatternFill(
                start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
            )  # Красный

        sheet[f"A{current_row}"].fill = fill
        sheet[f"B{current_row}"].fill = fill

        # Автоширина колонок
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 20

    def _create_dds_details_sheet(self, sheet, pnl_data: dict[str, Any]):
        """Создание листа с детализацией по платформам"""
        sheet["A1"] = "ДЕТАЛИЗАЦИЯ ПО ПЛАТФОРМАМ"
        sheet["A1"].font = Font(size=14, bold=True)

        current_row = 3

        # Wildberries детализация
        if pnl_data["wb"]["revenue"] > 0:
            sheet[f"A{current_row}"] = "WILDBERRIES"
            sheet[f"A{current_row}"].font = Font(size=12, bold=True)
            current_row += 1

            wb_data = pnl_data["wb"]

            details = [
                ("Общий объем заказов", wb_data.get("orders_revenue", 0)),
                ("Доставлено товаров", wb_data["revenue"]),
                ("Количество операций", wb_data["units"]),
                ("Комиссия + эквайринг", -wb_data["commission"]),
                ("Логистика + хранение", -wb_data.get("logistics_costs", 0)),
                ("Рекламные расходы", -wb_data.get("advertising_costs", 0)),
                ("К перечислению", wb_data.get("final_revenue", wb_data["revenue"])),
                ("Чистая прибыль", wb_data.get("profit", 0)),
            ]

            for desc, value in details:
                sheet[f"A{current_row}"] = desc
                sheet[f"B{current_row}"] = value
                sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
                current_row += 1

            current_row += 1

        # Ozon детализация
        if pnl_data["ozon"]["revenue"] > 0:
            sheet[f"A{current_row}"] = "OZON"
            sheet[f"A{current_row}"].font = Font(size=12, bold=True)
            current_row += 1

            ozon_data = pnl_data["ozon"]

            details = [
                ("Общий объем заказов", ozon_data.get("orders_revenue", 0)),
                ("Доставлено товаров", ozon_data["revenue"]),
                ("Количество операций", ozon_data["units"]),
                ("Комиссия", -ozon_data["commission"]),
                ("Реклама", -ozon_data.get("advertising_costs", 0)),
                ("Промо-акции", -ozon_data.get("promo_costs", 0)),
                ("Возвраты", -ozon_data.get("returns_cost", 0)),
                ("Логистика", -ozon_data.get("logistics_costs", 0)),
                ("Чистая прибыль", ozon_data.get("profit", 0)),
            ]

            for desc, value in details:
                sheet[f"A{current_row}"] = desc
                sheet[f"B{current_row}"] = value
                sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
                current_row += 1

        # Автоширина колонок
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 20

    async def generate_pnl_excel_report(
        self, date_from: str, date_to: str, cost_data_file: str | None = None
    ) -> str:
        """Генерация Excel таблицы P&L (прибыли и убытки)

        Args:
            date_from: Дата начала периода
            date_to: Дата окончания периода
            cost_data_file: Путь к файлу с данными о себестоимости (опционально)

        Returns:
            str: Путь к созданному Excel файлу

        """
        try:
            logger.info(f"Генерация P&L Excel отчета за период {date_from} - {date_to}")

            # Получаем финансовые данные
            pnl_data = await self.reports.calculate_real_pnl(date_from, date_to)

            # Загружаем данные о себестоимости если есть
            cost_data = None
            if cost_data_file and os.path.exists(cost_data_file):
                with open(cost_data_file, encoding="utf-8") as f:
                    cost_data = json.load(f)

            # Создаем Excel файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"PnL_report_{date_from}_{date_to}_{timestamp}.xlsx"
            filepath = f"/root/sovani_bot/excel_reports/{filename}"

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Создаем лист P&L
            pnl_sheet = wb.create_sheet("P&L отчет")
            self._create_pnl_sheet(pnl_sheet, pnl_data, date_from, date_to, cost_data)

            # Создаем лист с маржинальностью по SKU (если есть данные о себестоимости)
            if cost_data:
                sku_sheet = wb.create_sheet("Маржинальность SKU")
                self._create_sku_profitability_sheet(sku_sheet, pnl_data, cost_data)

            # Создаем лист сравнения с предыдущим периодом
            comparison_sheet = wb.create_sheet("Сравнение периодов")
            await self._create_comparison_sheet(comparison_sheet, pnl_data, date_from, date_to)

            wb.save(filepath)

            logger.info(f"P&L Excel отчет создан: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Ошибка создания P&L Excel отчета: {e}")
            raise

    def _create_pnl_sheet(
        self,
        sheet,
        pnl_data: dict[str, Any],
        date_from: str,
        date_to: str,
        cost_data: dict | None = None,
    ):
        """Создание основного листа P&L"""
        # Заголовок
        sheet["A1"] = "ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ (P&L)"
        sheet["A2"] = f"Период: {date_from} - {date_to}"
        sheet["A3"] = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        # Стили
        title_font = Font(size=14, bold=True)
        header_font = Font(size=11, bold=True)

        sheet["A1"].font = title_font
        sheet["A2"].font = header_font

        current_row = 6

        # ВЫРУЧКА
        total_revenue = pnl_data["total"]["revenue"]
        sheet[f"A{current_row}"] = "ВЫРУЧКА"
        sheet[f"B{current_row}"] = total_revenue
        sheet[f"C{current_row}"] = "100.0%"

        sheet[f"A{current_row}"].font = header_font
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        sheet[f"B{current_row}"].font = header_font
        current_row += 2

        # СЕБЕСТОИМОСТЬ
        total_cogs = pnl_data["total"]["cogs"]
        if cost_data and cost_data.get("sku_costs"):
            # Используем данные из шаблона себестоимости
            total_cogs = self._calculate_cogs_from_template(pnl_data, cost_data)

        cogs_percent = (total_cogs / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "Себестоимость товаров"
        sheet[f"B{current_row}"] = -total_cogs
        sheet[f"C{current_row}"] = f"-{cogs_percent:.1f}%"
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # ВАЛОВАЯ ПРИБЫЛЬ
        gross_profit = total_revenue - total_cogs
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "ВАЛОВАЯ ПРИБЫЛЬ"
        sheet[f"B{current_row}"] = gross_profit
        sheet[f"C{current_row}"] = f"{gross_margin:.1f}%"
        sheet[f"A{current_row}"].font = header_font
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        sheet[f"B{current_row}"].font = header_font
        current_row += 2

        # ОПЕРАЦИОННЫЕ РАСХОДЫ
        sheet[f"A{current_row}"] = "ОПЕРАЦИОННЫЕ РАСХОДЫ"
        sheet[f"A{current_row}"].font = header_font
        current_row += 1

        # Комиссии маркетплейсов
        total_commission = pnl_data["total"]["commission"]
        comm_percent = (total_commission / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "Комиссии маркетплейсов"
        sheet[f"B{current_row}"] = -total_commission
        sheet[f"C{current_row}"] = f"-{comm_percent:.1f}%"
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Реклама
        total_advertising = pnl_data["total"]["advertising"]
        ads_percent = (total_advertising / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "Рекламные расходы"
        sheet[f"B{current_row}"] = -total_advertising
        sheet[f"C{current_row}"] = f"-{ads_percent:.1f}%"
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 1

        # Прочие операционные расходы
        total_opex = pnl_data["total"]["opex"]
        opex_percent = (total_opex / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "Прочие операционные расходы"
        sheet[f"B{current_row}"] = -total_opex
        sheet[f"C{current_row}"] = f"-{opex_percent:.1f}%"
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"
        current_row += 2

        # ЧИСТАЯ ПРИБЫЛЬ
        net_profit = pnl_data["total"]["net_profit"]
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        sheet[f"A{current_row}"] = "ЧИСТАЯ ПРИБЫЛЬ"
        sheet[f"B{current_row}"] = net_profit
        sheet[f"C{current_row}"] = f"{net_margin:.1f}%"

        # Стилизация итоговой строки
        sheet[f"A{current_row}"].font = Font(size=12, bold=True)
        sheet[f"B{current_row}"].font = Font(size=12, bold=True)
        sheet[f"C{current_row}"].font = Font(size=12, bold=True)
        sheet[f"B{current_row}"].number_format = "#,##0.00 ₽"

        # Цвет фона
        if net_profit > 0:
            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        for col in ["A", "B", "C"]:
            sheet[f"{col}{current_row}"].fill = fill

        # Заголовки колонок
        sheet["A5"] = "Показатель"
        sheet["B5"] = "Сумма"
        sheet["C5"] = "% от выручки"

        for col in ["A", "B", "C"]:
            sheet[f"{col}5"].font = header_font

        # Автоширина колонок
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 15

    def _calculate_cogs_from_template(
        self, pnl_data: dict[str, Any], cost_data: dict[str, Any]
    ) -> float:
        """Расчет себестоимости на основе данных из шаблона"""
        # Пока возвращаем базовую себестоимость
        # TODO: Интегрировать с детальными данными продаж по SKU
        return pnl_data["total"]["cogs"]

    def _create_sku_profitability_sheet(
        self, sheet, pnl_data: dict[str, Any], cost_data: dict[str, Any]
    ):
        """Создание листа с анализом маржинальности по SKU"""
        sheet["A1"] = "АНАЛИЗ МАРЖИНАЛЬНОСТИ ПО SKU"
        sheet["A1"].font = Font(size=14, bold=True)

        # Заголовки таблицы
        headers = ["SKU", "Платформа", "Продано шт", "Выручка", "Себестоимость", "Маржа", "Маржа %"]
        for i, header in enumerate(headers, 1):
            sheet.cell(row=3, column=i, value=header)
            sheet.cell(row=3, column=i).font = Font(bold=True)

        # TODO: Добавить детальный анализ по SKU когда будут доступны данные продаж по товарам
        sheet["A5"] = (
            "Детальные данные по SKU будут доступны после интеграции с продажами по товарам"
        )

    async def _create_comparison_sheet(
        self, sheet, pnl_data: dict[str, Any], date_from: str, date_to: str
    ):
        """Создание листа сравнения с предыдущим периодом"""
        sheet["A1"] = "СРАВНЕНИЕ С ПРЕДЫДУЩИМ ПЕРИОДОМ"
        sheet["A1"].font = Font(size=14, bold=True)

        # Рассчитываем даты предыдущего периода
        current_start = datetime.strptime(date_from, "%Y-%m-%d")
        current_end = datetime.strptime(date_to, "%Y-%m-%d")
        period_days = (current_end - current_start).days + 1

        prev_end = current_start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days - 1)

        try:
            # Получаем данные предыдущего периода
            prev_pnl = await self.reports.calculate_real_pnl(
                prev_start.strftime("%Y-%m-%d"), prev_end.strftime("%Y-%m-%d")
            )

            # Создаем таблицу сравнения
            headers = [
                "Показатель",
                "Текущий период",
                "Предыдущий период",
                "Изменение",
                "Изменение %",
            ]
            for i, header in enumerate(headers, 1):
                sheet.cell(row=3, column=i, value=header)
                sheet.cell(row=3, column=i).font = Font(bold=True)

            # Данные для сравнения
            comparisons = [
                ("Выручка", pnl_data["total"]["revenue"], prev_pnl["total"]["revenue"]),
                ("Себестоимость", pnl_data["total"]["cogs"], prev_pnl["total"]["cogs"]),
                (
                    "Валовая прибыль",
                    pnl_data["total"]["gross_profit"],
                    prev_pnl["total"]["gross_profit"],
                ),
                ("Комиссии МП", pnl_data["total"]["commission"], prev_pnl["total"]["commission"]),
                ("Реклама", pnl_data["total"]["advertising"], prev_pnl["total"]["advertising"]),
                (
                    "Чистая прибыль",
                    pnl_data["total"]["net_profit"],
                    prev_pnl["total"]["net_profit"],
                ),
            ]

            row = 4
            for metric, current, previous in comparisons:
                change = current - previous
                change_percent = (change / previous * 100) if previous != 0 else 0

                sheet.cell(row=row, column=1, value=metric)
                sheet.cell(row=row, column=2, value=current).number_format = "#,##0.00 ₽"
                sheet.cell(row=row, column=3, value=previous).number_format = "#,##0.00 ₽"
                sheet.cell(row=row, column=4, value=change).number_format = "#,##0.00 ₽"
                sheet.cell(row=row, column=5, value=f"{change_percent:.1f}%")

                # Цвет для изменений
                if change > 0:
                    sheet.cell(row=row, column=4).font = Font(color="00FF00")  # Зеленый
                    sheet.cell(row=row, column=5).font = Font(color="00FF00")
                elif change < 0:
                    sheet.cell(row=row, column=4).font = Font(color="FF0000")  # Красный
                    sheet.cell(row=row, column=5).font = Font(color="FF0000")

                row += 1

        except Exception as e:
            logger.warning(f"Не удалось получить данные предыдущего периода: {e}")
            sheet["A5"] = "Данные предыдущего периода недоступны"

        # Автоширина колонок
        for i in range(1, 6):
            sheet.column_dimensions[chr(64 + i)].width = 20


# Глобальный экземпляр для использования в боте
excel_generator = ExcelReportGenerator()


async def test_excel_generator():
    """Тестирование генератора Excel отчетов"""
    try:
        # Тестируем DDS отчет
        dds_file = await excel_generator.generate_dds_excel_report("2024-09-01", "2024-09-07")
        print(f"DDS отчет создан: {dds_file}")

        # Тестируем P&L отчет
        pnl_file = await excel_generator.generate_pnl_excel_report("2024-09-01", "2024-09-07")
        print(f"P&L отчет создан: {pnl_file}")

    except Exception as e:
        print(f"Ошибка тестирования: {e}")


if __name__ == "__main__":
    asyncio.run(test_excel_generator())
