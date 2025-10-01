#!/usr/bin/env python3
"""SoVAni Bot - Telegram-бот для автоматизации работы продавца на Wildberries и Ozon

ОСНОВНЫЕ ФУНКЦИИ:
- Автоматическая обработка отзывов и вопросов с ChatGPT
- Генерация финансовых отчетов (P&L, выручка, ROI)
- Управление себестоимостью через Excel шаблоны
- Мониторинг складских остатков
- API интеграция с WB (6 токенов) и Ozon (2 токена)

АРХИТЕКТУРА:
- Framework: aiogram 2.25.1 (асинхронный Telegram Bot)
- Database: SQLite (sovani_bot.db)
- APIs: Wildberries API + Ozon API + OpenAI ChatGPT
- Authentication: JWT подписи для WB API
- Scheduler: APScheduler для автоматических задач

АВТОР: SoVAni Team
ВЕРСИЯ: 2.0 (Сентябрь 2025)
"""

import asyncio
import logging
from datetime import datetime

from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.types import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
)
from aiogram.utils import executor
from aiogram.utils.callback_data import CallbackData
from apscheduler.schedulers.asyncio import AsyncIOScheduler

import http_async
from ai_reply import generate_question_reply, generate_review_reply
from api_clients_main import (
    get_new_questions,
    get_new_reviews,
    post_answer_feedback,
    post_answer_question,
)
from api_monitor import api_monitor
from auto_reviews_processor import auto_processor

# NEW: pydantic-settings based configuration with validation
from app.core.config import config as Config, get_settings

from cost_template_generator import CostTemplateGenerator
from db import (
    get_question,
    get_review,
    init_db,
    mark_question_answered,
    mark_review_answered,
    save_question,
    save_review,
)

# Excel отчеты и интеграция с себестоимостью
from excel_bot_integration import (
    generate_cost_summary_for_bot,
    generate_enhanced_cumulative_report,
    generate_enhanced_financial_report,
    handle_cost_file_upload,
    handle_cost_template_upload,
    handle_dds_excel_export,
    handle_pnl_excel_export,
)
from handlers.api_client import range_preset, split_long
from handlers.inventory import repl_recommendations, stock_snapshot
from handlers.reports import dds_text, pnl_text, romi_text
from handlers.reviews import reviews_new_last24

# ЗАМЕНЕНО НА РЕАЛЬНУЮ СИСТЕМУ ОТЧЕТНОСТИ БЕЗ ФЕЙКОВ!
from real_data_reports import generate_cumulative_financial_report, generate_real_financial_report

# Система отзывов с ChatGPT
from reviews_bot_handlers import setup_reviews_handlers
from wb_excel_processor import wb_excel_processor

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("sovani_bot.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Validate configuration on startup (fail-fast)
try:
    settings = get_settings()
    logger.info("✅ Configuration validated successfully")
    logger.info(f"   - Telegram token: {settings.telegram_token[:10]}...")
    logger.info(f"   - Manager chat ID: {settings.manager_chat_id}")
    logger.info(f"   - OpenAI model: {settings.openai_model}")
    logger.info(f"   - Database: {settings.database_url}")
except RuntimeError as e:
    logger.error(f"❌ Configuration validation failed: {e}")
    raise

# Инициализация бота
bot = Bot(token=Config.TELEGRAM_TOKEN, parse_mode=types.ParseMode.HTML)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)
scheduler = AsyncIOScheduler()

# Callback data для кнопок
answer_callback = CallbackData("answer", "type", "id")

# Memory storage for user last report type
user_last = {}  # {user_id: 'pnl'|'dds'|'romi'}


# Простое меню для выбора периода финансового отчета
def get_financial_report_menu():
    """Меню для выбора периода финансового отчета"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("📅 7 дней", callback_data="fin_report:7"),
        InlineKeyboardButton("📅 30 дней", callback_data="fin_report:30"),
    )
    kb.add(
        InlineKeyboardButton("📅 Вчера", callback_data="fin_report:yesterday"),
        InlineKeyboardButton("📅 Сегодня", callback_data="fin_report:today"),
    )
    return kb


# Меню для нарастающего итога
def get_cumulative_report_menu():
    """Меню для выбора периода нарастающего итога"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("📊 7 дней", callback_data="cumulative:7"),
        InlineKeyboardButton("📊 30 дней", callback_data="cumulative:30"),
    )
    kb.add(
        InlineKeyboardButton("📊 60 дней", callback_data="cumulative:60"),
        InlineKeyboardButton("📊 90 дней", callback_data="cumulative:90"),
    )
    return kb


def split_message(text):
    """Разделение длинных сообщений на части для Telegram

    Telegram имеет ограничение на длину сообщения (4096 символов).
    Эта функция разбивает длинный текст на части по 3500 символов,
    стараясь делить по переносам строк для сохранения читабельности.

    Args:
        text (str): Исходный текст для разделения

    Returns:
        list[str]: Список частей сообщения

    """
    parts = []
    limit = 3500  # Безопасный лимит с запасом до 4096
    while len(text) > limit:
        # Ищем ближайший перенос строки до лимита
        cut = text.rfind("\n", 0, limit)
        cut = cut if cut != -1 else limit  # Если нет переноса, режем по лимиту
        parts.append(text[:cut])
        text = text[cut:]
    parts.append(text)
    return parts


# ===============================
# ГЛАВНОЕ МЕНЮ И НАВИГАЦИЯ
# ===============================


def get_main_menu():
    """Создание главного меню Telegram бота

    Главное меню содержит основные разделы функциональности:
    - Отчеты по платформам (WB/Ozon раздельно)
    - Управление данными (загрузка WB, себестоимость)
    - Работа с отзывами (автоматизация через ChatGPT)
    - Мониторинг системы (API статус)

    Returns:
        ReplyKeyboardMarkup: Клавиатура главного меню

    """
    print("[DEBUG] get_main_menu() вызвана! Создаю НОВОЕ меню с разделением WB/Ozon")
    logging.info("get_main_menu() called - creating NEW menu with WB/Ozon separation")

    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

    # Раздельные отчеты для WB и Ozon (основная функциональность)
    keyboard.add(
        KeyboardButton("🟣 Отчеты WB"),  # Wildberries финансовые отчеты
        KeyboardButton("🟠 Отчеты Ozon"),  # Ozon финансовые отчеты
    )

    # Управление данными
    keyboard.add(
        KeyboardButton("📋 Загрузка данных WB"),  # Excel импорт/экспорт для WB
        KeyboardButton("💰 Себестоимость"),  # Управление COGS через Excel
    )

    # Автоматизация и мониторинг
    keyboard.add(
        KeyboardButton("⭐ Управление отзывами"),  # ChatGPT автоответы
        KeyboardButton("🔍 API статус"),  # Проверка работоспособности API
    )

    # Справочная информация
    keyboard.add(KeyboardButton("📋 Помощь"))
    return keyboard


# Меню отчетов WB
def get_wb_reports_menu():
    """Меню отчетов Wildberries"""
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("📊 Финансовый отчет", callback_data="wb_financial"),
        InlineKeyboardButton("📈 Нарастающий итог", callback_data="wb_cumulative"),
    )
    keyboard.add(
        InlineKeyboardButton("📋 Товарные остатки", callback_data="wb_stock"),
        InlineKeyboardButton("🎯 Топ товаров", callback_data="wb_top_sku"),
    )
    keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="main_menu"))
    return keyboard


# Меню отчетов Ozon
def get_ozon_reports_menu():
    """Меню отчетов Ozon"""
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("📊 Финансовый отчет", callback_data="ozon_financial"),
        InlineKeyboardButton("📈 Нарастающий итог", callback_data="ozon_cumulative"),
    )
    keyboard.add(
        InlineKeyboardButton("📋 Товарные остатки", callback_data="ozon_stock"),
        InlineKeyboardButton("🎯 Аналитика продаж", callback_data="ozon_analytics"),
    )
    keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="main_menu"))
    return keyboard


# Меню загрузки данных WB
def get_wb_upload_menu():
    """Меню загрузки данных WB"""
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("📤 Загрузить отчет о продажах", callback_data="wb_upload_sales"),
        InlineKeyboardButton("📤 Загрузить отчет о заказах", callback_data="wb_upload_orders"),
        InlineKeyboardButton("📤 Загрузить финансовый отчет", callback_data="wb_upload_finance"),
    )
    keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="main_menu"))
    return keyboard


# Меню себестоимости
def get_cost_menu():
    """Меню управления себестоимостью"""
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        InlineKeyboardButton("📋 Шаблон себестоимости", callback_data="cost_template"),
        InlineKeyboardButton("📊 Сводка себестоимости", callback_data="cost_summary"),
    )
    keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="main_menu"))
    return keyboard


# Команды быстрого доступа
@dp.message_handler(commands=["reports"])
async def reports_command(message: types.Message):
    """Команда быстрого финансового отчета"""
    await financial_report_handler(message)


@dp.message_handler(commands=["api_status"])
async def api_status_command(message: types.Message):
    """Команда проверки статуса API"""
    await api_status_handler(message)


@dp.message_handler(commands=["help"])
async def help_command(message: types.Message):
    """Команда помощи"""
    await help_handler(message)


# === НОВЫЕ ОБРАБОТЧИКИ ГЛАВНОГО МЕНЮ ===


@dp.message_handler(text="🟣 Отчеты WB")
async def wb_reports_handler(message: types.Message):
    """Показать меню отчетов Wildberries"""
    await message.answer(
        "🟣 <b>Отчеты Wildberries</b>\n\n" "📊 Выберите тип отчета:",
        reply_markup=get_wb_reports_menu(),
        parse_mode="HTML",
    )


@dp.message_handler(text="🟠 Отчеты Ozon")
async def ozon_reports_handler(message: types.Message):
    """Показать меню отчетов Ozon"""
    await message.answer(
        "🟠 <b>Отчеты Ozon</b>\n\n" "📊 Выберите тип отчета:",
        reply_markup=get_ozon_reports_menu(),
        parse_mode="HTML",
    )


@dp.message_handler(text="📋 Загрузка данных WB")
async def wb_upload_handler(message: types.Message):
    """Показать меню загрузки данных WB"""
    await message.answer(
        "📋 <b>Загрузка данных Wildberries</b>\n\n"
        "📤 Загрузите Excel файлы с отчетами WB для анализа более поздних периодов:\n\n"
        "💡 <i>Эта функция позволяет загружать отчеты, скачанные с личного кабинета WB, "
        "для анализа периодов старше 176 дней</i>",
        reply_markup=get_wb_upload_menu(),
        parse_mode="HTML",
    )


@dp.message_handler(text="💰 Себестоимость")
async def cost_handler(message: types.Message):
    """Показать меню управления себестоимостью"""
    await message.answer(
        "💰 <b>Управление себестоимостью</b>\n\n" "📊 Выберите действие:",
        reply_markup=get_cost_menu(),
        parse_mode="HTML",
    )


@dp.message_handler(text="💰 Шаблон себестоимости")
async def cost_template_handler(message: types.Message):
    """Генерация и отправка шаблона себестоимости"""
    status_msg = await message.answer("💰 Генерирую шаблон себестоимости...")

    try:
        generator = CostTemplateGenerator()
        template_path = await generator.generate_cost_template()

        await status_msg.edit_text("💰 Шаблон готов! Отправляю файл...")

        await message.answer_document(
            types.InputFile(
                template_path,
                filename=f"cost_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            ),
            caption="💰 <b>Шаблон для загрузки себестоимости</b>\n\n"
            "📋 <b>Инструкция:</b>\n"
            "1. Заполните себестоимость для каждого SKU\n"
            "2. Укажите переменные расходы (за единицу)\n"
            "3. Внесите постоянные расходы (в месяц)\n"
            "4. Отправьте заполненный файл обратно в бот\n\n"
            "📊 После обработки вы получите анализ прибыльности по каждому товару",
            parse_mode="HTML",
        )

        await status_msg.delete()

    except Exception as e:
        logger.error(f"Ошибка генерации шаблона себестоимости: {e}")
        await status_msg.edit_text(f"❌ Ошибка генерации шаблона: {str(e)[:100]}")


@dp.message_handler(text="🔍 API статус")
async def api_status_handler(message: types.Message):
    """Проверить статус API"""
    status_msg = await message.answer("🔍 Проверяю статус API...")

    try:
        report = await api_monitor.generate_status_report()
        await status_msg.edit_text(report, parse_mode="HTML")
    except Exception as e:
        logger.error(f"Ошибка проверки API: {e}")
        await status_msg.edit_text(f"❌ Ошибка: {str(e)[:100]}")


@dp.message_handler(text="⭐ Управление отзывами")
async def reviews_management_handler(message: types.Message):
    """Управление отзывами через кнопку меню"""
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("🆕 Новые отзывы", callback_data="reviews_new"),
        InlineKeyboardButton("📋 ВСЕ неотвеченные", callback_data="reviews_all"),
        InlineKeyboardButton("⚙️ Настройки отзывов", callback_data="reviews_settings"),
    )

    await message.answer(
        "⭐ <b>Управление отзывами</b>\n\n"
        "🆕 <b>Новые отзывы</b> - последние поступившие отзывы\n"
        "📋 <b>ВСЕ неотвеченные</b> - обработка всех отзывов без ответа (первое подключение)\n"
        "⚙️ <b>Настройки</b> - управление автоответами\n\n"
        "Выберите действие:",
        parse_mode="HTML",
        reply_markup=keyboard,
    )


@dp.callback_query_handler(lambda c: c.data == "reviews_new")
async def handle_reviews_new_callback(callback_query: types.CallbackQuery):
    """Обработка новых отзывов"""
    from reviews_bot_handlers import ReviewsBotHandlers

    await callback_query.message.delete()
    await ReviewsBotHandlers.handle_reviews_command(callback_query.message)
    await callback_query.answer()


@dp.callback_query_handler(lambda c: c.data == "reviews_all")
async def handle_reviews_all_callback(callback_query: types.CallbackQuery):
    """Обработка всех неотвеченных отзывов"""
    from reviews_bot_handlers import ReviewsBotHandlers

    await callback_query.message.delete()
    await ReviewsBotHandlers.handle_all_unanswered_reviews_command(callback_query.message)
    await callback_query.answer()


@dp.callback_query_handler(lambda c: c.data == "reviews_settings")
async def handle_reviews_settings_callback(callback_query: types.CallbackQuery):
    """Настройки отзывов"""
    await callback_query.message.edit_text(
        "⚙️ <b>Настройки отзывов</b>\n\n"
        "🔄 Автоматическая проверка: 2 раза в день (9:00, 18:00)\n"
        "✅ Автоответ на отзывы 4-5⭐\n"
        "⚠️ Ручная проверка отзывов 1-3⭐\n"
        "🤖 ChatGPT интеграция: включена\n\n"
        "Используйте /reviews для новых отзывов\n"
        "Используйте /all_reviews для всех неотвеченных",
        parse_mode="HTML",
    )
    await callback_query.answer()


@dp.message_handler(text="📋 Помощь")
async def help_handler(message: types.Message):
    """Показать помощь"""
    help_text = """
📋 <b>Помощь SoVAni Bot</b>

<b>Доступные функции:</b>
📊 <b>Финансовый отчет</b> - полный финансовый анализ
📈 <b>Нарастающий итог</b> - анализ за период
💰 <b>Шаблон себестоимости</b> - генерация Excel для учета затрат
⭐ <b>Управление отзывами</b> - обработка отзывов с ChatGPT
🔍 <b>API статус</b> - проверка подключений
📋 <b>Помощь</b> - эта справка

<b>Команды:</b>
/start - главное меню
/reports - быстрый отчет
/reviews - новые отзывы
/all_reviews - ВСЕ неотвеченные отзывы
/api_status - статус API

💡 <i>Все функции работают с реальными данными из WB и Ozon API</i>
    """
    await message.answer(help_text, parse_mode="HTML")


# Обработчик кнопок нарастающего итога
@dp.callback_query_handler(lambda c: c.data and c.data.startswith("cumulative:"))
async def cumulative_callback_handler(call: types.CallbackQuery):
    """Обработка выбора периода для нарастающего итога"""
    period = call.data.split(":")[1]

    await bot.answer_callback_query(call.id)
    await bot.send_message(call.from_user.id, f"📈 Генерирую нарастающий итог за {period}...")

    try:
        # Используем существующую логику отчетов
        from_date, to_date = range_preset(period)
        report = await pnl_text(from_date, to_date, "day")

        for part in split_long(report):
            await bot.send_message(call.from_user.id, part, parse_mode="HTML")
    except Exception as e:
        logger.error(f"Ошибка нарастающего итога: {e}")
        await bot.send_message(call.from_user.id, f"❌ Ошибка: {str(e)[:100]}")


# === NEW REAL HANDLERS ===


def presets_kb():
    kb = InlineKeyboardMarkup(row_width=3)
    kb.add(
        InlineKeyboardButton("Сегодня", callback_data="rng:today"),
        InlineKeyboardButton("Вчера", callback_data="rng:yesterday"),
        InlineKeyboardButton("7 дней", callback_data="rng:7d"),
        InlineKeyboardButton("30 дней", callback_data="rng:30d"),
    )
    return kb


@dp.message_handler(commands=["start"])
async def start_command(message: types.Message):
    """Команда /start"""
    print(f"[DEBUG] /start команда от пользователя {message.from_user.id}")
    logging.info(f"START command received from user {message.from_user.id}")
    welcome_text = """
🎆 <b>Добро пожаловать в SoVAni Bot!</b>

📊 Получайте реальные финансовые отчеты из WB и Ozon
🔍 Мониторьте статус API подключений
📈 Анализируйте динамику продаж

Используйте кнопки меню ниже для начала работы.
    """
    await message.answer(welcome_text, parse_mode="HTML", reply_markup=get_main_menu())


@dp.message_handler(commands=["api_status", "status"])
async def cmd_api_status(message: types.Message):
    """Проверка статуса API интеграций"""
    await message.answer("🔍 Проверяю статус API интеграций...")

    try:
        report = await api_monitor.generate_status_report()

        # Разбиваем длинный отчет на части если нужно
        for part in split_long(report):
            await message.answer(part, parse_mode="HTML")

        # Дополнительно отправляем краткую сводку для удобства
        summary = api_monitor.get_status_summary()

        await message.answer(
            f"📋 <b>Краткая сводка:</b>\n"
            f"{summary['message']}\n\n"
            f"⏰ Последняя проверка: {summary['last_check'].strftime('%H:%M:%S')}"
        )

    except Exception as e:
        logger.error(f"Ошибка проверки статуса API: {e}")
        await message.answer(f"❌ Ошибка проверки статуса API: {e}")


# Обработчики inline-кнопок меню


# Обработчики меню отчетов
@dp.callback_query_handler(text="report_full")
async def full_report_callback(callback_query: types.CallbackQuery):
    """Полный отчет"""
    await callback_query.answer()
    await callback_query.message.edit_text("📊 Генерирую полный отчет...")

    try:
        # РЕАЛЬНАЯ СИСТЕМА ОТЧЕТНОСТИ БЕЗ ФЕЙКОВ!
        report = await generate_real_financial_report()
        await callback_query.message.edit_text(report, parse_mode="HTML")
    except Exception as e:
        await callback_query.message.edit_text(f"❌ Ошибка: {str(e)[:200]}")


@dp.callback_query_handler(text="report_cumulative")
async def cumulative_callback(callback_query: types.CallbackQuery):
    """Обработчик нарастающего итога P&L"""
    await callback_query.answer()

    status_msg = await callback_query.message.edit_text("🔄 Генерирую нарастающий итог...")

    try:
        # Создаем инлайн-клавиатуру с выбором периода
        keyboard = InlineKeyboardMarkup(row_width=2)
        keyboard.add(
            InlineKeyboardButton("7 дней", callback_data="cumulative_7"),
            InlineKeyboardButton("30 дней", callback_data="cumulative_30"),
        )
        keyboard.add(
            InlineKeyboardButton("60 дней", callback_data="cumulative_60"),
            InlineKeyboardButton("90 дней", callback_data="cumulative_90"),
        )
        keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="reports_menu"))

        await status_msg.edit_text(
            "📊 <b>Нарастающий итог P&L</b>\n\nВыберите период для анализа:",
            parse_mode="HTML",
            reply_markup=keyboard,
        )
    except Exception as e:
        await status_msg.edit_text(f"❌ Ошибка: {e!s}")
        logger.error(f"Ошибка меню нарастающего итога: {e}")


@dp.callback_query_handler(lambda c: c.data.startswith("cumulative_"))
async def cumulative_period_callback(callback_query: types.CallbackQuery):
    """Обработчик выбора периода для нарастающего итога"""
    await callback_query.answer()

    days = int(callback_query.data.split("_")[1])
    status_msg = await callback_query.message.edit_text(
        f"🔄 Генерирую нарастающий итог за {days} дней..."
    )

    try:
        report = await generate_cumulative_financial_report(days)

        # Добавляем кнопку "Назад"
        keyboard = InlineKeyboardMarkup()
        keyboard.add(
            InlineKeyboardButton("⬅️ Назад к выбору периода", callback_data="report_cumulative")
        )
        keyboard.add(InlineKeyboardButton("🔄 Обновить", callback_data=f"cumulative_{days}"))

        await status_msg.edit_text(report, parse_mode="HTML", reply_markup=keyboard)
    except Exception as e:
        await status_msg.edit_text(f"❌ Ошибка: {e!s}")
        logger.error(f"Ошибка генерации нарастающего итога за {days} дней: {e}")


@dp.callback_query_handler(text="report_wb_extended")
async def wb_extended_report_callback(callback_query: types.CallbackQuery):
    """Расширенный отчет Wildberries через async API"""
    await callback_query.answer("📢 Загружаю расширенные данные WB...")
    await callback_query.message.edit_text("📢 Генерирую расширенный отчет Wildberries...")

    try:
        from datetime import date, timedelta

        # Получаем данные за последние 7 дней через async API
        date_to = date.today()
        date_from = date_to - timedelta(days=7)

        # Собираем данные параллельно
        pnl_task = http_async.get_json(
            "/finance/pnl",
            params={
                "platform": "wb",
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        ads_task = http_async.get_json(
            "/ads/overview",
            params={
                "platform": "wb",
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        inventory_task = http_async.get_json("/inventory/stock_snapshot", params={"platform": "wb"})

        # Ожидаем все результаты
        pnl_data, ads_data, inventory_data = await asyncio.gather(
            pnl_task, ads_task, inventory_task, return_exceptions=True
        )

        # Формируем отчёт
        report_parts = ["📢 <b>Расширенный отчёт Wildberries</b>\n"]
        missing_parts = []

        # P&L данные
        if isinstance(pnl_data, dict) and pnl_data.get("platforms", {}).get("wb"):
            wb_data = pnl_data["platforms"]["wb"]
            report_parts.append(f"💰 <b>Выручка:</b> {wb_data.get('revenue', '—')} ₽")
            report_parts.append(f"🏪 <b>Комиссия МП:</b> {wb_data.get('mp_commission', '—')} ₽")
            report_parts.append(f"🚚 <b>Логистика:</b> {wb_data.get('logistics', '—')} ₽")
        else:
            missing_parts.append("Финансовые данные недоступны")

        # Рекламные данные
        if isinstance(ads_data, dict) and ads_data.get("platforms", {}).get("wb"):
            wb_ads = ads_data["platforms"]["wb"]
            report_parts.append(f"📈 <b>Реклама:</b> {wb_ads.get('ads_cost', '—')} ₽")
        else:
            missing_parts.append("Рекламные данные недоступны")

        # Остатки
        if isinstance(inventory_data, dict) and inventory_data.get("platforms", {}).get("wb"):
            wb_stock = inventory_data["platforms"]["wb"]
            report_parts.append(f"📦 <b>Остатки:</b> {wb_stock.get('total_quantity', '—')} ед.")
        else:
            missing_parts.append("Данные остатков недоступны")

        # Добавляем примечания
        if missing_parts:
            report_parts.append("\n⚠️ <b>Примечания:</b>")
            for part in missing_parts:
                report_parts.append(f"• {part}")

        report_parts.append(f"\n📅 <i>Период: {date_from} - {date_to}</i>")

        report = "\n".join(report_parts)
        await callback_query.message.edit_text(report, parse_mode="HTML")

    except Exception as e:
        logger.error(f"Ошибка расширенного WB отчета: {e}")
        await callback_query.message.edit_text(
            f"❌ Ошибка получения данных: {str(e)[:200]}", parse_mode="HTML"
        )


# Обработчики настроек
@dp.callback_query_handler(text="settings_tokens")
async def tokens_settings_callback(callback_query: types.CallbackQuery):
    """Показать информацию о токенах"""
    await callback_query.answer()
    text = f"""🔑 <b>API Токены</b>

🟣 <b>Wildberries:</b>
• Отзывы: {'✅ Настроен' if Config.WB_API_TOKEN else '❌ Не настроен'}
• Отчеты: {'✅ Настроен' if Config.WB_REPORTS_TOKEN else '❌ Не настроен'}

🔵 <b>Ozon:</b>
• Client-ID: {'✅ Настроен' if Config.OZON_CLIENT_ID else '❌ Не настроен'}
• API-Key: {'✅ Настроен' if Config.OZON_API_KEY else '❌ Не настроен'}

🤖 <b>OpenAI:</b>
• API-Key: {'✅ Настроен' if Config.OPENAI_API_KEY else '❌ Не настроен'}

ℹ️ Для изменения токенов обратитесь к администратору"""

    back_button = InlineKeyboardMarkup().add(
        InlineKeyboardButton("🔙 Назад к настройкам", callback_data="back_settings")
    )
    await callback_query.message.edit_text(text, reply_markup=back_button)


@dp.callback_query_handler(text="back_settings")
async def back_to_settings(callback_query: types.CallbackQuery):
    """Возврат в настройки"""
    await callback_query.answer()
    await settings_handler(callback_query.message)


# Date range callback handler
@dp.callback_query_handler(lambda c: c.data and c.data.startswith("rng:"))
async def cb_range(call: types.CallbackQuery):
    preset = call.data.split(":", 1)[1]
    uid = call.from_user.id
    last = user_last.get(uid, "pnl")
    f, t = preset_range(preset)
    await bot.answer_callback_query(call.id)
    await bot.send_message(
        uid, f"Выбран период: {f}–{t}. Собираю {last.upper()}…", parse_mode="HTML"
    )
    try:
        if last == "pnl":
            text = report_pnl(f, t, "day")
        elif last == "dds":
            text = report_dds(f, t)
        else:
            text = report_romi(f, t)
        for part in split_message(text):
            await bot.send_message(uid, part, parse_mode="HTML")
    except Exception as e:
        await bot.send_message(uid, f"❌ Ошибка: {e}", parse_mode="HTML")


@dp.callback_query_handler(answer_callback.filter())
async def handle_answer_callback(callback_query: types.CallbackQuery, callback_data: dict):
    """Обработка нажатия кнопки 'Ответить'"""
    await callback_query.answer()

    answer_type = callback_data["type"]  # 'review' или 'question'
    item_id = callback_data["id"]

    logger.info(f"🔔 Получен callback: type={answer_type}, id={item_id}")

    try:
        if answer_type == "review":
            logger.info(f"📝 Ищем отзыв в базе: {item_id}")
            review = get_review(item_id)

            if not review:
                logger.error(f"❌ Отзыв {item_id} не найден в базе данных")
                await callback_query.message.edit_reply_markup(reply_markup=None)
                await callback_query.message.reply("❌ Отзыв не найден в базе данных")
                return

            logger.info(f"✅ Отзыв найден: {review['id']}, ответ: {review['answer'][:50]}...")

            # Отправляем ответ через API Wildberries
            logger.info("📤 Отправляем ответ через API...")
            success = await post_answer_feedback(review["id"], review["answer"])

            logger.info(f"📊 Результат отправки: {success}")

            if success:
                mark_review_answered(review["id"])
                # Обновляем сообщение согласно ТЗ - добавляем отметку в конце
                original_text = callback_query.message.text
                updated_text = original_text + "\n\n(✅ Ответ отправлен)"

                await callback_query.message.edit_text(
                    text=updated_text, parse_mode="HTML", reply_markup=None
                )
                logger.info("✅ Интерфейс обновлен: ответ отмечен как отправленный")
            else:
                logger.error(f"❌ Не удалось отправить ответ на отзыв {item_id}")
                await callback_query.message.reply(
                    "⚠️ Не удалось отправить ответ (ошибка API). Попробуйте позже."
                )

        elif answer_type == "question":
            question = get_question(item_id)
            if not question:
                await callback_query.message.edit_reply_markup(reply_markup=None)
                await callback_query.message.reply("❌ Вопрос не найден в базе данных")
                return

            # Отправляем ответ через API Wildberries
            success = await post_answer_question(question["id"], question["answer"])

            if success:
                mark_question_answered(question["id"])
                new_markup = InlineKeyboardMarkup().add(
                    InlineKeyboardButton("✅ Ответ отправлен", callback_data="sent")
                )
                await callback_query.message.edit_reply_markup(reply_markup=new_markup)
                await callback_query.message.reply("✅ Ответ на вопрос успешно отправлен!")
            else:
                await callback_query.message.reply(
                    "⚠️ Не удалось отправить ответ. Проверьте API настройки."
                )

    except Exception as e:
        logger.error(f"Ошибка при отправке ответа: {e}")
        await callback_query.message.reply(f"❌ Ошибка: {e!s}")


async def check_new_reviews_and_questions():
    """Проверка новых отзывов и вопросов (запускается по расписанию)"""
    logger.info("Начинаю проверку новых отзывов и вопросов...")

    try:
        # Проверяем новые отзывы
        new_reviews = await get_new_reviews()
        logger.info(f"Найдено {len(new_reviews)} новых отзывов")

        for review in new_reviews:
            # Генерируем ответ на отзыв
            reply = await generate_review_reply(review)

            # Сохраняем в базу данных
            save_review(
                {
                    "id": review["id"],
                    "sku": review.get("sku", "N/A"),
                    "text": review.get("text", ""),
                    "rating": review.get("rating", 0),
                    "has_media": review.get("has_media", False),
                    "answer": reply,
                    "date": review.get("date", datetime.now()),
                    "answered": False,
                }
            )

            # Формируем сообщение для Telegram согласно ТЗ
            stars = "⭐" * review.get("rating", 0)
            media_info = " [📷 Фото от клиента]" if review.get("has_media") else ""
            review_text = review.get("text") or "отзыв без текста"
            product_name = review.get("product_name", review.get("sku", "N/A"))

            # ТОЛЬКО РЕАЛЬНЫЕ ОТЗЫВЫ - БЕЗ ДЕМО РЕЖИМА!

            message_text = f"""📝 Новый отзыв на товар <b>{product_name}</b> – {stars}{media_info}:
"<i>{review_text}</i>"
<b>Предложенный ответ:</b> {reply}"""

            # Кнопка для отправки ответа согласно ТЗ
            keyboard = InlineKeyboardMarkup().add(
                InlineKeyboardButton(
                    "Ответить ✅", callback_data=answer_callback.new(type="review", id=review["id"])
                )
            )

            try:
                logger.info(
                    f"📤 Отправляем уведомление об отзыве {review['id']} в чат {Config.MANAGER_CHAT_ID}"
                )
                await bot.send_message(
                    Config.MANAGER_CHAT_ID, message_text, reply_markup=keyboard, parse_mode="HTML"
                )
                logger.info(f"✅ Уведомление об отзыве {review['id']} отправлено")
            except Exception as send_error:
                logger.error(
                    f"❌ Ошибка отправки уведомления об отзыве {review['id']}: {send_error}"
                )

        # Проверяем новые вопросы
        new_questions = await get_new_questions()
        logger.info(f"Найдено {len(new_questions)} новых вопросов")

        for question in new_questions:
            # Генерируем ответ на вопрос
            reply = await generate_question_reply(question)

            # Сохраняем в базу данных
            save_question(
                {
                    "id": question["id"],
                    "sku": question.get("sku", "N/A"),
                    "text": question.get("text", ""),
                    "answer": reply,
                    "date": question.get("date", datetime.now()),
                    "answered": False,
                }
            )

            # Формируем сообщение для Telegram
            message_text = f"""❓ <b>Вопрос от покупателя по {question.get('sku', 'N/A')}:</b>
"<i>{question.get('text', 'N/A')}</i>"

<b>Предложенный ответ:</b> {reply}"""

            # Кнопка для отправки ответа
            keyboard = InlineKeyboardMarkup().add(
                InlineKeyboardButton(
                    "Ответить ✅",
                    callback_data=answer_callback.new(type="question", id=question["id"]),
                )
            )

            await bot.send_message(
                Config.MANAGER_CHAT_ID, message_text, reply_markup=keyboard, parse_mode="HTML"
            )

    except Exception as e:
        logger.error(f"Ошибка при проверке отзывов и вопросов: {e}")
        await bot.send_message(
            Config.MANAGER_CHAT_ID, f"⚠️ Ошибка при проверке новых отзывов и вопросов: {e!s}"
        )


async def on_startup(dp):
    """Действия при запуску бота"""
    logger.info("Запуск SoVAni Bot...")

    # Инициализация базы данных
    init_db()
    logger.info("База данных инициализирована")

    # Инициализация HTTP клиента
    await http_async.init_http_client()
    logger.info("HTTP клиент инициализирован")

    # Настройка команд меню бота
    # Новое минимальное меню команд
    commands = [
        BotCommand("start", "🎆 Главное меню"),
        BotCommand("reports", "📊 Финансовый отчет"),
        BotCommand("reviews", "⭐ Новые отзывы"),
        BotCommand("all_reviews", "📋 ВСЕ неотвеченные отзывы"),
        BotCommand("api_status", "🔍 API статус"),
        BotCommand("help", "📋 Помощь"),
    ]
    await bot.set_my_commands(commands)
    logger.info("Команды меню настроены")

    # Настройка обработчиков отзывов
    await setup_reviews_handlers(dp)
    logger.info("Обработчики отзывов настроены")

    # Инициализация автопроцессора отзывов
    auto_processor.bot = bot
    auto_processor.set_admin_chat_id(Config.MANAGER_CHAT_ID)

    # Настройка планировщика задач
    scheduler.add_job(
        check_new_reviews_and_questions,
        "cron",
        hour=6,  # Каждый день в 06:00
        minute=0,
        timezone="Europe/Moscow",
    )

    # Добавляем задачу автообработки отзывов (каждые 4 часа)
    scheduler.add_job(
        auto_processor._process_reviews_cycle,
        "cron",
        hour="6,10,14,18,22",  # Каждые 4 часа: 06:00, 10:00, 14:00, 18:00, 22:00
        minute=0,
        timezone="Europe/Moscow",
    )

    scheduler.start()
    logger.info(
        "Планировщик задач запущен (проверка отзывов каждые 4 часа: 6:00, 10:00, 14:00, 18:00, 22:00)"
    )

    # Уведомление о запуске
    try:
        await bot.send_message(
            Config.MANAGER_CHAT_ID,
            "🤖 <b>SoVAni Bot запущен!</b>\n\n"
            "✅ Автоматическая проверка отзывов включена (каждый день в 06:00)\n"
            "🤖 Автообработка отзывов с ChatGPT (каждые 4 часа: 6:00, 10:00, 14:00, 18:00, 22:00)\n"
            "⭐ Команда /reviews для управления отзывами\n"
            "✅ Обработка команд активна\n"
            "✅ Анализ отчетов доступен\n"
            "🔄 Запускаю первичную обработку ВСЕХ неотвеченных отзывов...",
        )

        # Запускаем первичную обработку всех неотвеченных отзывов при старте
        await auto_processor.process_all_unanswered_reviews()
        logger.info("Первичная обработка всех неотвеченных отзывов завершена")

    except Exception as e:
        logger.error(f"Не удалось отправить уведомление о запуске: {e}")


async def on_shutdown(dp):
    """Действия при остановке бота"""
    logger.info("Остановка SoVAni Bot...")

    # Закрываем HTTP клиент
    await http_async.close_http_client()
    logger.info("HTTP клиент закрыт")

    try:
        scheduler.shutdown()
    except:
        pass


# == STRICT REALITY MODE ==
from handlers.rca import ads as _rca_ads
from handlers.rca import finance as _rca_finance
from handlers.rca import inventory as _rca_inventory
from handlers.rca import reviews as _rca_reviews


def _fmt_missing(miss: list):
    if not miss:
        return ""
    return "\n".join([f"• {m}" for m in miss])


@dp.message_handler(commands=["pnl_strict"])
async def pnl_real(message: types.Message):
    f, t = range_preset("7d")
    try:
        st = await _rca_finance(f, t)
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>Отчёт P&L не построен.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> откройте https://justbusiness.lol/finance и загрузите/включите недостающие данные."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = await pnl_text(f, t, "day")
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка P&L: {e}", parse_mode="HTML")


@dp.message_handler(commands=["dds_strict"])
async def dds_real(message: types.Message):
    f, t = range_preset("7d")
    try:
        st = await _rca_finance(f, t)
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>DDS не построен.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> загрузите Cashflow CSV и COGS/OPEX на https://justbusiness.lol/finance."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = dds_text(f, t)
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка DDS: {e}", parse_mode="HTML")


@dp.message_handler(commands=["romi_strict"])
async def romi_real(message: types.Message):
    f, t = range_preset("7d")
    try:
        st = await _rca_ads(f, t)
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>ROMI не построен.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> загрузите кампании и статистику рекламы на https://justbusiness.lol/ads."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = romi_text(f, t)
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка ROMI: {e}", parse_mode="HTML")


@dp.message_handler(commands=["stock_strict"])
async def stock_real(message: types.Message):
    try:
        st = await _rca_inventory()
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>Остатки не отображены.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> загрузите снепшот остатков и параметры поставок на https://justbusiness.lol/inventory."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = await stock_snapshot()
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка Остатков: {e}", parse_mode="HTML")


@dp.message_handler(commands=["repl_strict"])
async def repl_real(message: types.Message):
    try:
        st = await _rca_inventory()
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>Рекомендации не рассчитаны.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> добавьте параметры поставок и историю продаж на https://justbusiness.lol/inventory."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = await repl_recommendations()
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка Рекомендаций: {e}", parse_mode="HTML")


@dp.message_handler(commands=["reviews_strict"])
async def reviews_real(message: types.Message):
    try:
        st = await _rca_reviews()
        miss = st.get("missing", [])
        if miss:
            txt = (
                "<b>Отзывы недоступны.</b>\n<b>Причины:</b>\n"
                + _fmt_missing(miss)
                + "\n\n<b>Действия:</b> загрузите отзывы CSV и добавьте шаблоны ответов на https://justbusiness.lol/reviews."
            )
            for part in split_long(txt):
                await message.answer(part, parse_mode="HTML")
            return
        txt = reviews_new_last24()
        for part in split_long(txt):
            await message.answer(part, parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Ошибка Отзывов: {e}", parse_mode="HTML")


@dp.message_handler(commands=["diag_bot"])
async def diag_bot_handler(message: types.Message):
    """Диагностика бота (BOT-ASYNC-FIX)"""
    try:
        await message.answer("🔧 Выполняю диагностику бота...", parse_mode="HTML")

        # Проверка конфигурации
        import os
        from datetime import date, timedelta

        bot_token_ok = bool(os.getenv("BOT_TOKEN"))
        backend_url = http_async.get_backend_base_url()
        service_token_ok = bool(http_async.get_service_token())

        # Проверка health endpoints
        health_results = await http_async.health_check()

        # Проверка типового P&L запроса
        date_to = date.today()
        date_from = date_to - timedelta(days=7)
        sample_pnl = await http_async.get_json(
            "/finance/pnl",
            params={
                "platform": "all",
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        # Поиск проблемных паттернов в коде
        import glob

        problematic_patterns = 0
        try:
            for py_file in glob.glob("/root/sovani_bot/*.py"):
                with open(py_file) as f:
                    content = f.read()
                    if (
                        "asyncio.run(" in content
                        or "run_until_complete(" in content
                        or "nest_asyncio" in content
                    ):
                        problematic_patterns += 1
        except:
            pass  # Игнорируем ошибки файлового доступа

        # Формируем отчёт
        report = f"""🔧 <b>Диагностика SoVAni Bot</b>

<b>🤖 Конфигурация:</b>
• Bot Token: {'✅' if bot_token_ok else '❌'} {'настроен' if bot_token_ok else 'не настроен'}
• Backend URL: <code>{backend_url}</code>
• X-Service-Token: {'✅' if service_token_ok else '❌'} {'настроен' if service_token_ok else 'не настроен'}

<b>🌐 Бэкенд:</b>
• Подключение: {'✅' if health_results['backend_reachable'] else '❌'} {'доступен' if health_results['backend_reachable'] else 'недоступен'}
• Health статус: {health_results.get('health_status', 'неизвестно')}
• Ops Health: {health_results.get('ops_health_status', 'неизвестно')}

<b>🔄 Async структура:</b>
• HTTP Client: httpx.AsyncClient
• Event Loop Nested Calls: {problematic_patterns} найдено
• Bot Async OK: {'yes' if problematic_patterns == 0 else 'no'}

<b>📊 Тестовый API вызов:</b>
• Sample PnL: {'OK' if sample_pnl else 'error'}
{f"  Data: {len(sample_pnl)} полей" if sample_pnl else "  Error: нет данных или ошибка связи"}

<b>🎯 Статус:</b>
{'✅ Все системы готовы к работе!' if (bot_token_ok and service_token_ok and health_results['backend_reachable'] and problematic_patterns == 0) else '⚠️ Есть проблемы, требующие внимания'}"""

        # Отправляем отчёт
        parts = split_message(report)
        for part in parts:
            await message.answer(part, parse_mode="HTML")

    except Exception as e:
        await message.answer(f"❌ Ошибка диагностики: {e}", parse_mode="HTML")
        logger.error(f"Ошибка в diag_bot: {e}")


@dp.message_handler(commands=["debug_backend"])
async def debug_backend_handler(message: types.Message):
    """Последовательное тестирование бэкенд endpoints"""
    from datetime import date, timedelta

    await message.answer("🔧 Тестирую бэкенд endpoints...", parse_mode="HTML")

    # Задаем тестовые параметры
    today = date.today()
    date_from = (today - timedelta(days=3)).isoformat()
    date_to = today.isoformat()

    # Список endpoints для тестирования
    endpoints = [
        ("/health", {}),
        ("/live/finance/pnl_v2", {"platform": "wb", "date_from": date_from, "date_to": date_to}),
        ("/live/inventory/stock", {"platform": "wb"}),
        ("/live/ads/overview", {"platform": "wb", "date_from": date_from, "date_to": date_to}),
    ]

    results = []

    for endpoint, params in endpoints:
        try:
            import time

            start_time = time.time()

            result = await http_async.get_json(endpoint, params=params)

            duration = int((time.time() - start_time) * 1000)
            status = "✅ 200" if result else "❌ No data"
            response_preview = str(result)[:200] if result else "Empty response"

            results.append(
                {
                    "endpoint": endpoint,
                    "status": status,
                    "duration": duration,
                    "preview": response_preview,
                }
            )

        except Exception as e:
            import time

            duration = int((time.time() - start_time) * 1000)
            error_msg = http_async.format_error_message(e)

            results.append(
                {
                    "endpoint": endpoint,
                    "status": f"❌ {error_msg}",
                    "duration": duration,
                    "preview": str(e)[:100],
                }
            )

    # Формируем отчет
    report = ["🔧 <b>Backend Endpoints Test</b>\n"]

    for r in results:
        report.append(f"<b>{r['endpoint']}</b>")
        report.append(f"Status: {r['status']}")
        report.append(f"Duration: {r['duration']}ms")
        report.append(f"Preview: <code>{r['preview']}</code>\n")

    report_text = "\n".join(report)
    parts = split_message(report_text)

    for part in parts:
        await message.answer(part, parse_mode="HTML")


@dp.message_handler(commands=["diag_loop"])
async def diag_loop_handler(message: types.Message):
    """Проверка запрещенных async паттернов в боте"""
    await message.answer("🔍 Проверяю async паттерны в коде...", parse_mode="HTML")

    import glob

    # Запрещенные паттерны
    banned_patterns = [
        "asyncio.run(",
        "run_until_complete(",
        "nest_asyncio",
        "requests.get(",
        "requests.post(",
    ]

    violations = []

    try:
        # Проверяем все Python файлы в директории бота
        for py_file in glob.glob("/root/sovani_bot/*.py") + glob.glob(
            "/root/sovani_bot/**/*.py", recursive=True
        ):
            if ".bak" in py_file or "__pycache__" in py_file:
                continue

            try:
                with open(py_file, encoding="utf-8") as f:
                    lines = f.readlines()

                for line_num, line in enumerate(lines, 1):
                    for pattern in banned_patterns:
                        if pattern in line and not line.strip().startswith("#"):
                            violations.append(
                                {
                                    "file": py_file.replace("/root/sovani_bot/", ""),
                                    "line": line_num,
                                    "pattern": pattern,
                                    "code": line.strip()[:80],
                                }
                            )
            except Exception:
                continue  # Пропускаем файлы с ошибками чтения

    except Exception as e:
        await message.answer(f"❌ Ошибка сканирования: {e}", parse_mode="HTML")
        return

    # Формируем отчет
    if not violations:
        report = """✅ <b>Bot Async OK: yes</b>

🔍 <b>Проверено:</b>
• asyncio.run() - не найдено
• run_until_complete() - не найдено
• nest_asyncio - не найдено
• requests.get() - не найдено
• requests.post() - не найдено

✨ Все async паттерны корректны!"""
    else:
        report = ["❌ <b>Bot Async OK: no</b>"]
        report.append(f"<b>Найдено нарушений:</b> {len(violations)}\n")

        for v in violations[:10]:  # Показываем первые 10
            report.append(f"<b>{v['file']}:{v['line']}</b>")
            report.append(f"Pattern: <code>{v['pattern']}</code>")
            report.append(f"Code: <code>{v['code']}</code>\n")

        if len(violations) > 10:
            report.append(f"... и ещё {len(violations) - 10} нарушений")

    report_text = "\n".join(report) if isinstance(report, list) else report
    parts = split_message(report_text)

    for part in parts:
        await message.answer(part, parse_mode="HTML")


# ====== UNIFIED REPORTING COMMANDS (PROMPT-BOT-REPORT-V2) ======


def get_period_keyboard():
    """Клавиатура для выбора периода"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("📅 Сегодня", callback_data="period:today"),
        InlineKeyboardButton("📅 Вчера", callback_data="period:yesterday"),
    )
    kb.add(
        InlineKeyboardButton("📅 7 дней", callback_data="period:7d"),
        InlineKeyboardButton("📅 30 дней", callback_data="period:30d"),
    )
    return kb


def get_platform_keyboard(command):
    """Клавиатура для выбора платформы"""
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🟣 Wildberries", callback_data=f"{command}:wb"),
        InlineKeyboardButton("🔵 Ozon", callback_data=f"{command}:ozon"),
    )
    kb.add(InlineKeyboardButton("📊 Все платформы", callback_data=f"{command}:all"))
    return kb


async def format_currency(amount):
    """Форматирование валют с обработкой ошибок API"""
    if amount is None or amount == 0:
        return "—"
    try:
        return f"{amount:,.0f} ₽".replace(",", " ")
    except:
        return "—"


async def format_percentage(value):
    """Форматирование процентов"""
    if value is None:
        return "—"
    try:
        return f"{value:.1f}%"
    except:
        return "—"


@dp.message_handler(commands=["pnl"])
async def pnl_command(message: types.Message):
    """Команда P&L - прибыли и убытки"""
    await message.answer(
        "💰 <b>P&L - Прибыли и убытки</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("pnl"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("pnl:"))
async def pnl_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для P&L"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "💰 <b>P&L - Прибыли и убытки</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    # Сохраняем выбранную платформу
    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "pnl", "platform": platform})


@dp.message_handler(commands=["romi"])
async def romi_command(message: types.Message):
    """Команда ROMI - возврат на маркетинговые инвестиции"""
    await message.answer(
        "📈 <b>ROMI - Возврат на маркетинговые инвестиции</b>\n\n"
        "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("romi"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("romi:"))
async def romi_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для ROMI"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "📈 <b>ROMI - Возврат на маркетинговые инвестиции</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "romi", "platform": platform})


@dp.message_handler(commands=["dds"])
async def dds_command(message: types.Message):
    """Команда DDS - детализация продаж по дням"""
    await message.answer(
        "📊 <b>DDS - Daily Detailed Sales</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("dds"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("dds:"))
async def dds_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для DDS"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "📊 <b>DDS - Daily Detailed Sales</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "dds", "platform": platform})


@dp.message_handler(commands=["stock"])
async def stock_command(message: types.Message):
    """Команда остатки на складах"""
    await message.answer(
        "📦 <b>Остатки на складах</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("stock"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("stock:"))
async def stock_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для остатков"""
    platform = callback.data.split(":")[1]

    # Для остатков период не нужен - показываем текущие данные
    await callback.message.edit_text("📦 <b>Загружаю данные об остатках...</b>", parse_mode="HTML")

    await show_stock_report(callback.message, platform)


@dp.message_handler(commands=["reviews"])
async def reviews_command(message: types.Message):
    """Команда статистика отзывов"""
    await message.answer(
        "⭐ <b>Статистика отзывов</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("reviews"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("reviews:"))
async def reviews_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для отзывов"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "⭐ <b>Статистика отзывов</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "reviews", "platform": platform})


@dp.callback_query_handler(lambda c: c.data.startswith("period:"))
async def period_callback(callback: types.CallbackQuery):
    """Обработка выбора периода"""
    from date_picker import DatePicker, date_range_manager, get_calendar_for_date_selection

    period = callback.data.split(":")[1]

    # Обработка кастомного периода
    if period == "custom":
        # Получаем контекст из текущей сессии пользователя
        user_id = callback.from_user.id
        selection = date_range_manager.get_selection(user_id)
        context = selection.get("context") if selection else None

        await callback.message.edit_text(
            "📅 <b>Выбор периода</b>\n\n" "Выберите начальную дату:",
            reply_markup=get_calendar_for_date_selection(selection_type="from", context=context),
            parse_mode="HTML",
        )
        await callback.answer()
        return

    # Обработка предустановленных периодов
    try:
        date_from_str, date_to_str = DatePicker.parse_predefined_period(period)
        period_description = DatePicker.format_period_description(date_from_str, date_to_str)

        # Получаем контекст пользователя
        selection = date_range_manager.get_selection(callback.from_user.id)
        if not selection:
            await callback.answer("❌ Ошибка: контекст команды потерян")
            return

        context = selection.get("context", "financial_report")

        # Генерируем соответствующий отчет
        await generate_report_for_period(
            callback.message, context, date_from_str, date_to_str, period_description
        )

        # Очищаем выбор пользователя
        date_range_manager.clear_selection(callback.from_user.id)

    except Exception as e:
        logger.error(f"Ошибка обработки периода {period}: {e}")
        await callback.answer("❌ Ошибка обработки периода")
        return

    await callback.answer()


# Обработчики календаря
@dp.callback_query_handler(lambda c: c.data.startswith("calendar:"))
async def calendar_callback(callback: types.CallbackQuery):
    """Обработка календарных действий"""
    from datetime import datetime

    from date_picker import DatePicker, date_range_manager, get_calendar_for_date_selection

    data_parts = callback.data.split(":")
    action = data_parts[1]

    if action == "ignore":
        await callback.answer()
        return

    elif action == "cancel":
        date_range_manager.clear_selection(callback.from_user.id)
        await callback.message.edit_text("❌ Выбор периода отменен")
        await callback.answer()
        return

    elif action == "quick":
        # Возврат к быстрым периодам
        from date_picker import get_enhanced_period_menu

        await callback.message.edit_text(
            "📊 <b>Финансовый отчет</b>\n\n" "📅 Выберите период для анализа:",
            reply_markup=get_enhanced_period_menu(),
            parse_mode="HTML",
        )
        await callback.answer()
        return

    elif action in ["from", "to"]:
        # Навигация по месяцам
        year, month = int(data_parts[2]), int(data_parts[3])
        # Получаем контекст из текущей сессии пользователя
        user_id = callback.from_user.id
        selection = date_range_manager.get_selection(user_id)
        context = selection.get("context") if selection else None

        await callback.message.edit_reply_markup(
            reply_markup=get_calendar_for_date_selection(year, month, action, context)
        )
        await callback.answer()
        return

    elif action == "select":
        # Выбор даты
        selection_type = data_parts[2]  # from или to
        year, month, day = int(data_parts[3]), int(data_parts[4]), int(data_parts[5])
        selected_date = f"{year:04d}-{month:02d}-{day:02d}"

        user_id = callback.from_user.id
        selection = date_range_manager.get_selection(user_id)

        if not selection:
            await callback.answer("❌ Ошибка: сессия выбора потеряна")
            return

        if selection_type == "from":
            # Выбрана начальная дата
            date_range_manager.set_date_from(user_id, selected_date)

            formatted_date = datetime.strptime(selected_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            # Получаем контекст для календаря конечной даты
            selection = date_range_manager.get_selection(user_id)
            context = selection.get("context") if selection else None

            await callback.message.edit_text(
                f"📅 <b>Выбор периода</b>\n\n"
                f"✅ Начальная дата: {formatted_date}\n"
                f"📅 Выберите конечную дату:",
                reply_markup=get_calendar_for_date_selection(selection_type="to", context=context),
                parse_mode="HTML",
            )

        elif selection_type == "to":
            # Выбрана конечная дата
            date_range_manager.set_date_to(user_id, selected_date)

            selection = date_range_manager.get_selection(user_id)
            date_from = selection["date_from"]
            date_to = selected_date
            context = selection.get("context", "financial_report")

            # Валидация диапазона с учетом контекста маркетплейса
            is_valid, error_msg = DatePicker.validate_date_range(date_from, date_to, context)
            if not is_valid:
                await callback.message.edit_text(
                    f"❌ <b>Ошибка выбора дат</b>\n\n" f"{error_msg}\n\n" f"Попробуйте снова:",
                    reply_markup=get_calendar_for_date_selection(
                        selection_type="from", context=context
                    ),
                    parse_mode="HTML",
                )
                await callback.answer()
                return

            # Генерируем отчет
            period_description = DatePicker.format_period_description(date_from, date_to)
            context = selection.get("context", "financial_report")

            await generate_report_for_period(
                callback.message, context, date_from, date_to, period_description
            )
            date_range_manager.clear_selection(user_id)

        await callback.answer()


async def generate_report_for_period(
    message, context: str, date_from: str, date_to: str, period_description: str
):
    """Генерация отчета для выбранного периода"""
    # Обновляем сообщение о загрузке
    await message.edit_text(
        f"📊 <b>Генерирую отчет...</b>\n\n"
        f"📅 Период: {period_description}\n"
        f"🔄 Загрузка данных...",
        parse_mode="HTML",
    )

    try:
        if context == "financial_report":
            # Расширенный финансовый отчет с кнопками экспорта
            report, markup = await generate_enhanced_financial_report(date_from, date_to, message)

        elif context == "cumulative_report":
            # Расширенный нарастающий итог с кнопками экспорта
            from datetime import datetime

            days = (
                datetime.strptime(date_to, "%Y-%m-%d") - datetime.strptime(date_from, "%Y-%m-%d")
            ).days + 1
            report, markup = await generate_enhanced_cumulative_report(days)

        elif context == "wb_financial":
            # WB финансовый отчет
            from excel_bot_integration import generate_wb_financial_report

            report, markup = await generate_wb_financial_report(date_from, date_to, message)

        elif context == "ozon_financial":
            # Ozon финансовый отчет
            from excel_bot_integration import generate_ozon_financial_report

            report, markup = await generate_ozon_financial_report(date_from, date_to, message)

        else:
            report = f"❌ Неизвестный тип отчета: {context}"
            markup = None

        # Разбиваем длинные сообщения
        parts = split_message(report)

        # Отправляем первую часть как редактирование с кнопками
        await message.edit_text(
            parts[0],
            parse_mode="HTML",
            disable_web_page_preview=True,
            reply_markup=markup if len(parts) == 1 else None,
        )

        # Остальные части как новые сообщения
        for i, part in enumerate(parts[1:]):
            # Добавляем кнопки только к последней части
            part_markup = markup if i == len(parts) - 2 else None
            await message.answer(
                part, parse_mode="HTML", disable_web_page_preview=True, reply_markup=part_markup
            )

    except Exception as e:
        logger.error(f"Ошибка генерации отчета для {context}: {e}")
        await message.edit_text(
            f"❌ <b>Ошибка генерации отчета</b>\n\n"
            f"📅 Период: {period_description}\n"
            f"❗ {str(e)[:200]}",
            parse_mode="HTML",
        )


async def show_pnl_report(message, platform, date_from, date_to, period_name):
    """Показать отчёт P&L через реальную систему данных"""
    try:
        # НОВАЯ СИСТЕМА: Используем реальные данные вместо API endpoint
        from real_data_reports import RealDataFinancialReports

        real_reports = RealDataFinancialReports()

        # Конвертируем platform для совместимости
        if platform == "all":
            platform_filter = "both"
        elif platform == "wb":
            platform_filter = "wb"
        elif platform == "ozon":
            platform_filter = "ozon"
        else:
            platform_filter = "both"

        # Получаем данные через нашу реальную систему
        pnl_data = await real_reports.calculate_real_pnl(
            date_from.strftime("%Y-%m-%d"),
            date_to.strftime("%Y-%m-%d"),
            platform_filter=platform_filter,
        )

        if not pnl_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных P&L</b>\n\n"
                "API недоступен или вернул пустой результат",
                parse_mode="HTML",
            )
            return

        # НОВАЯ СИСТЕМА: Используем format_real_pnl_report для форматирования
        base_report = real_reports.format_real_pnl_report(pnl_data)

        # Добавляем информацию о платформе в заголовок
        platform_name = {"wb": "🟣 Wildberries", "ozon": "🔵 Ozon", "both": "📊 Все платформы"}.get(
            platform_filter, "📊 Все платформы"
        )

        # Добавляем информацию о производительности
        processing_time = pnl_data.get("processing_time", 0)
        parallelized = pnl_data.get("parallelized", False)
        chunked = pnl_data.get("chunked", False)

        performance_info = f"\n\n⚡ Обработка: {processing_time:.1f}с"
        if parallelized:
            performance_info += " (параллельно"
            if chunked:
                performance_info += ", chunked"
            performance_info += ")"

        report = f"💰 <b>P&L - Прибыли и убытки</b>\n{platform_name}\n📅 <b>{period_name}</b>\n\n{base_report}{performance_info}"

        # Добавляем кнопки экспорта
        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "📤 Экспорт CSV", callback_data=f"export_pnl:{platform}:{date_from}:{date_to}"
            ),
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_pnl:{platform}:{date_from}:{date_to}"
            ),
        )

        await message.edit_text(report, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке P&L</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка P&L: {e}")


async def show_romi_report(message, platform, date_from, date_to, period_name):
    """Показать отчёт ROMI"""
    try:
        # Запрос данных ROMI
        romi_data = await http_async.get_json(
            "/ads/overview",
            params={
                "platform": platform,
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        if not romi_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных ROMI</b>\n\n"
                "API недоступен или вернул пустой результат",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        ads_spend = romi_data.get("ads_spend", 0)
        ads_revenue = romi_data.get("ads_revenue", 0)
        romi_percent = romi_data.get("romi_percent", 0)
        clicks = romi_data.get("clicks", 0)
        impressions = romi_data.get("impressions", 0)
        ctr = romi_data.get("ctr", 0)

        report = f"""📈 <b>ROMI - Возврат на маркетинговые инвестиции</b>

{platform_emoji} <b>{platform_name}</b>
📅 <b>{period_name}</b>

💸 <b>Потрачено на рекламу:</b> {await format_currency(ads_spend)}
💵 <b>Выручка от рекламы:</b> {await format_currency(ads_revenue)}
📈 <b>ROMI:</b> {await format_percentage(romi_percent)}

👀 <b>Показы:</b> {impressions:,}
👆 <b>Клики:</b> {clicks:,}
💡 <b>CTR:</b> {await format_percentage(ctr)}

<i>Данные обновлены: {datetime.now().strftime('%H:%M')}</i>"""

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_romi:{platform}:{date_from}:{date_to}"
            )
        )

        await message.edit_text(report, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке ROMI</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка ROMI: {e}")


async def show_dds_report(message, platform, date_from, date_to, period_name):
    """Показать отчёт DDS"""
    try:
        dds_data = await http_async.get_json(
            "/finance/dds",
            params={
                "platform": platform,
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        if not dds_data or not dds_data.get("daily_sales"):
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных DDS</b>\n\n"
                "API недоступен или нет данных за период",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        daily_sales = dds_data["daily_sales"]
        total_revenue = sum(day.get("revenue", 0) for day in daily_sales)
        total_orders = sum(day.get("orders", 0) for day in daily_sales)

        report = f"""📊 <b>DDS - Daily Detailed Sales</b>

{platform_emoji} <b>{platform_name}</b>
📅 <b>{period_name}</b>

💰 <b>Общая выручка:</b> {await format_currency(total_revenue)}
🛒 <b>Всего заказов:</b> {total_orders:,}

📈 <b>По дням:</b>\n"""

        for day in daily_sales[-10:]:  # Показываем последние 10 дней
            day_date = day.get("date", "")
            day_revenue = day.get("revenue", 0)
            day_orders = day.get("orders", 0)

            report += f"• {day_date}: {await format_currency(day_revenue)} ({day_orders} зак.)\n"

        if len(daily_sales) > 10:
            report += f"\n<i>... и ещё {len(daily_sales) - 10} дней</i>\n"

        report += f"\n<i>Данные обновлены: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_dds:{platform}:{date_from}:{date_to}"
            )
        )

        await message.edit_text(report, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке DDS</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка DDS: {e}")


async def show_stock_report(message, platform):
    """Показать отчёт по остаткам"""
    try:
        # Используем новый live endpoint
        stock_data = await http_async.get_json(
            "/live/inventory/stock", params={"platform": platform}
        )

        if not stock_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных остатков</b>\n\n" "API недоступен или нет данных",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        # Новый формат ответа от live endpoint
        if platform == "all":
            # Комбинированный ответ
            wb_stocks = stock_data.get("wb", [])
            ozon_stocks = stock_data.get("ozon", [])
            total_items = len(wb_stocks) + len(ozon_stocks)

            text = "📦 <b>Остатки на складах</b>\n\n"

            # Wildberries секция
            if wb_stocks:
                text += "🟣 <b>Wildberries</b>\n"
                wb_total_qty = sum(item.get("quantity", 0) for item in wb_stocks)
                text += f"📊 Товаров: {len(wb_stocks)}\n"
                text += f"📦 Общее кол-во: {wb_total_qty:,}\n\n"
            else:
                text += "🟣 <b>Wildberries</b>\n📊 Нет данных\n\n"

            # Ozon секция
            if ozon_stocks:
                text += "🔵 <b>Ozon</b>\n"
                ozon_total_qty = sum(item.get("quantity", 0) for item in ozon_stocks)
                text += f"📊 Товаров: {len(ozon_stocks)}\n"
                text += f"📦 Общее кол-во: {ozon_total_qty:,}\n\n"
            else:
                text += "🔵 <b>Ozon</b>\n📊 Нет данных\n\n"

            text += "📊 <b>ИТОГО</b>\n"
            text += f"📊 Всего товаров: {total_items}\n"
            text += f"📦 Общее кол-во: {sum(item.get('quantity', 0) for item in wb_stocks + ozon_stocks):,}"

        else:
            # Одна платформа
            stocks = stock_data if isinstance(stock_data, list) else stock_data.get("data", [])
            total_items = len(stocks)
            total_quantity = sum(item.get("quantity", 0) for item in stocks)
            total_value = sum(item.get("value", 0) for item in stocks)

            text = f"""📦 <b>Остатки на складах</b>

{platform_emoji} <b>{platform_name}</b>
🕐 <b>На данный момент</b>

📊 <b>Всего товаров:</b> {total_items}
📦 <b>Общее количество:</b> {total_quantity:,} шт."""

            if total_value > 0:
                text += f"\n💰 <b>Общая стоимость:</b> {await format_currency(total_value)}"

            if stocks:
                text += "\n\n🔝 <b>Топ товары по остаткам:</b>\n"

                # Сортируем по количеству и показываем топ-10
                top_stocks = sorted(stocks, key=lambda x: x.get("quantity", 0), reverse=True)[:10]

                for i, item in enumerate(top_stocks, 1):
                    name = item.get("name", item.get("vendor_code", "Без названия"))[:30]
                    quantity = item.get("quantity", 0)
                    price = item.get("price", 0)

                    text += f"{i}. <b>{name}</b>\n"
                    if price > 0:
                        text += f"   📦 {quantity} шт. × {await format_currency(price)}\n"
                    else:
                        text += f"   📦 {quantity} шт.\n"

            text += f"\n<i>Данные обновлены: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(InlineKeyboardButton("🔄 Обновить", callback_data=f"refresh_stock:{platform}"))

        await message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке остатков</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка остатков: {e}")


async def show_reviews_report(message, platform, date_from, date_to, period_name):
    """Показать отчёт по отзывам"""
    try:
        reviews_data = await http_async.get_json(
            "/reviews/stats",
            params={
                "platform": platform,
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
            },
        )

        if not reviews_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки статистики отзывов</b>\n\n"
                "API недоступен или вернул пустой результат",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        total_reviews = reviews_data.get("total_reviews", 0)
        avg_rating = reviews_data.get("avg_rating", 0)
        five_star = reviews_data.get("five_star", 0)
        four_star = reviews_data.get("four_star", 0)
        three_star = reviews_data.get("three_star", 0)
        two_star = reviews_data.get("two_star", 0)
        one_star = reviews_data.get("one_star", 0)

        report = f"""⭐ <b>Статистика отзывов</b>

{platform_emoji} <b>{platform_name}</b>
📅 <b>{period_name}</b>

💬 <b>Всего отзывов:</b> {total_reviews}
⭐ <b>Средняя оценка:</b> {avg_rating:.1f}

📊 <b>Распределение оценок:</b>
⭐⭐⭐⭐⭐ {five_star} ({five_star/max(total_reviews,1)*100:.0f}%)
⭐⭐⭐⭐ {four_star} ({four_star/max(total_reviews,1)*100:.0f}%)
⭐⭐⭐ {three_star} ({three_star/max(total_reviews,1)*100:.0f}%)
⭐⭐ {two_star} ({two_star/max(total_reviews,1)*100:.0f}%)
⭐ {one_star} ({one_star/max(total_reviews,1)*100:.0f}%)

<i>Данные обновлены: {datetime.now().strftime('%H:%M')}</i>"""

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_reviews:{platform}:{date_from}:{date_to}"
            )
        )

        await message.edit_text(report, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке отзывов</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка отзывов: {e}")


@dp.message_handler(content_types=["document"])
async def handle_document_upload(message: types.Message):
    """Обработка загрузки файлов (шаблонов себестоимости)"""
    try:
        document = message.document
        if not document:
            return

        # Проверяем тип файла
        if not document.file_name.endswith((".xlsx", ".xls")):
            await message.reply("❌ Поддерживаются только Excel файлы (.xlsx, .xls)")
            return

        # Проверяем что это шаблон себестоимости
        if "cost_template" not in document.file_name.lower():
            await message.reply("❌ Загружайте только заполненные шаблоны себестоимости")
            return

        status_msg = await message.reply("📄 Обрабатываю загруженный шаблон...")

        # Скачиваем файл
        file_info = await bot.get_file(document.file_id)
        downloaded_file = await bot.download_file(file_info.file_path)

        # Сохраняем файл
        import os

        os.makedirs("/root/sovani_bot/uploaded_templates", exist_ok=True)
        file_path = f"/root/sovani_bot/uploaded_templates/uploaded_{document.file_name}"
        with open(file_path, "wb") as f:
            f.write(downloaded_file.read())

        # Обрабатываем загруженный шаблон
        generator = CostTemplateGenerator()
        processed_data = await generator.process_filled_template(file_path)

        # Сохраняем обработанные данные
        saved_path = await generator.save_cost_data(processed_data)

        # Формируем отчет об обработке
        summary = processed_data["summary"]
        response_text = f"""
💰 <b>Шаблон себестоимости обработан!</b>

📊 <b>Результаты обработки:</b>
• SKU с себестоимостью: {summary['total_sku_with_costs']}
• Переменные расходы: {summary['total_variable_costs']}
• Постоянные расходы: {summary['total_fixed_costs']}
• Сумма постоянных расходов: {summary.get('total_fixed_monthly', 0):,.0f} ₽/мес

📁 Данные сохранены: <code>{saved_path}</code>

🔄 Теперь эти данные можно использовать для расчета прибыльности каждого SKU
        """

        await status_msg.edit_text(response_text, parse_mode="HTML")

    except Exception as e:
        logger.error(f"Ошибка обработки шаблона себестоимости: {e}")
        await message.reply(f"❌ Ошибка обработки файла: {str(e)[:100]}")


@dp.message_handler(commands=["export_pnl"])
async def export_pnl_command(message: types.Message):
    """Экспорт P&L в CSV"""
    await message.answer(
        "📤 <b>Экспорт P&L в CSV</b>\n\n" "Выберите платформу:",
        reply_markup=get_platform_keyboard("export_pnl"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("export_pnl:"))
async def export_pnl_callback(callback: types.CallbackQuery):
    """Экспорт P&L данных в CSV"""
    try:
        parts = callback.data.split(":")
        platform = parts[1]

        if len(parts) >= 4:
            # Экспорт из существующего отчёта
            date_from = parts[2]
            date_to = parts[3]
        else:
            # Новый экспорт - используем последние 30 дней
            from datetime import date, timedelta

            today = date.today()
            date_from = (today - timedelta(days=30)).isoformat()
            date_to = today.isoformat()

        await callback.message.edit_text("📤 <b>Генерирую CSV файл...</b>", parse_mode="HTML")

        # Получаем данные P&L
        pnl_data = await http_async.get_json(
            "/finance/pnl",
            params={
                "platform": platform,
                "date_from": date_from,
                "date_to": date_to,
                "detailed": "true",  # Детализированные данные для CSV
            },
        )

        if not pnl_data:
            await callback.message.edit_text(
                "❌ <b>Ошибка экспорта</b>\n\nДанные недоступны", parse_mode="HTML"
            )
            return

        # Генерируем CSV
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        writer.writerow(
            ["Дата", "Платформа", "Выручка", "Расходы", "Реклама", "Прибыль", "Маржа_%"]
        )

        # Данные (если есть детализация по дням)
        if "daily_data" in pnl_data:
            for day in pnl_data["daily_data"]:
                writer.writerow(
                    [
                        day.get("date", ""),
                        platform.upper(),
                        day.get("revenue", 0),
                        day.get("costs", 0),
                        day.get("ads_spend", 0),
                        day.get("profit", 0),
                        day.get("margin_percent", 0),
                    ]
                )
        else:
            # Итоговые данные
            writer.writerow(
                [
                    f"{date_from} - {date_to}",
                    platform.upper(),
                    pnl_data.get("revenue", 0),
                    pnl_data.get("costs", 0),
                    pnl_data.get("ads_spend", 0),
                    pnl_data.get("profit", 0),
                    pnl_data.get("margin_percent", 0),
                ]
            )

        # Отправляем файл
        csv_content = output.getvalue()
        output.close()

        platform_name = {"wb": "wildberries", "ozon": "ozon", "all": "all_platforms"}[platform]
        filename = f"pnl_{platform_name}_{date_from}_{date_to}.csv"

        from io import BytesIO

        file_bytes = BytesIO(csv_content.encode("utf-8"))
        file_bytes.name = filename

        await bot.send_document(
            callback.message.chat.id,
            file_bytes,
            caption=f"📤 <b>P&L экспорт</b>\n\n"
            f"Платформа: {platform.upper()}\n"
            f"Период: {date_from} — {date_to}",
            parse_mode="HTML",
        )

        await callback.message.delete()

    except Exception as e:
        await callback.message.edit_text(
            f"❌ <b>Ошибка экспорта</b>\n\n<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка экспорта P&L: {e}")


@dp.message_handler(commands=["status_live"])
async def status_live_command(message: types.Message):
    """Статус API и системы в реальном времени"""
    try:
        await message.answer("🔄 <b>Проверяю статус системы...</b>", parse_mode="HTML")

        # Проверяем все основные API endpoints параллельно
        health_task = http_async.health_check()

        # Простые запросы для проверки доступности
        from datetime import date

        today = date.today()

        pnl_test_task = http_async.get_json(
            "/finance/pnl",
            params={
                "platform": "all",
                "date_from": today.isoformat(),
                "date_to": today.isoformat(),
            },
        )

        stock_test_task = http_async.get_json("/inventory/stock", params={"platform": "all"})

        # Ждём результаты
        health_result, pnl_result, stock_result = await asyncio.gather(
            health_task, pnl_test_task, stock_test_task, return_exceptions=True
        )

        # Формируем отчёт
        report = f"""🚦 <b>Статус системы LIVE</b>

🕐 <b>Проверено:</b> {datetime.now().strftime('%H:%M:%S')}

🌐 <b>Backend Health:</b>"""

        if isinstance(health_result, dict):
            backend_status = "✅ Работает"
            db_status = "✅ ОК" if health_result.get("database") == "ok" else "❌ Ошибка"
        elif isinstance(health_result, Exception):
            backend_status = "❌ Недоступен"
            db_status = "❌ Недоступна"
        else:
            backend_status = "⚠️ Неизвестно"
            db_status = "⚠️ Неизвестно"

        report += f"\n• API: {backend_status}"
        report += f"\n• База данных: {db_status}"

        report += "\n\n📊 <b>API Endpoints:</b>"

        # P&L API
        if isinstance(pnl_result, dict):
            pnl_status = "✅ Работает"
        elif isinstance(pnl_result, Exception):
            pnl_status = "❌ Ошибка"
        else:
            pnl_status = "⚠️ Неизвестно"
        report += f"\n• P&L (/finance/pnl): {pnl_status}"

        # Stock API
        if isinstance(stock_result, dict):
            stock_status = "✅ Работает"
        elif isinstance(stock_result, Exception):
            stock_status = "❌ Ошибка"
        else:
            stock_status = "⚠️ Неизвестно"
        report += f"\n• Остатки (/inventory/stock): {stock_status}"

        # Общий статус
        all_ok = all(
            [
                isinstance(health_result, dict),
                isinstance(pnl_result, dict),
                isinstance(stock_result, dict),
            ]
        )

        if all_ok:
            overall_status = "✅ Все системы работают"
        else:
            overall_status = "⚠️ Обнаружены проблемы"

        report += f"\n\n🎯 <b>Общий статус:</b> {overall_status}"

        # Кнопка обновления
        kb = InlineKeyboardMarkup()
        kb.add(InlineKeyboardButton("🔄 Обновить статус", callback_data="refresh_status"))

        await message.answer(report, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.answer(
            f"❌ <b>Ошибка проверки статуса</b>\n\n<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка status_live: {e}")


@dp.callback_query_handler(lambda c: c.data == "refresh_status")
async def refresh_status_callback(callback: types.CallbackQuery):
    """Обновление статуса системы"""
    await callback.message.edit_text("🔄 <b>Обновляю статус системы...</b>", parse_mode="HTML")
    await status_live_command(callback.message)
    await callback.answer("✅ Статус обновлён")


# ====== NEW WB INTEGRATION COMMANDS (PROMPT-WB-FULL-IMPLEMENTATION) ======


@dp.message_handler(commands=["ads"])
async def ads_command(message: types.Message):
    """Команда статистики рекламы"""
    await message.answer(
        "📊 <b>Статистика рекламы</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("ads"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("ads:"))
async def ads_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для рекламы"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "📊 <b>Статистика рекламы</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "ads", "platform": platform})


@dp.message_handler(commands=["supplies"])
async def supplies_command(message: types.Message):
    """Команда отчёта поставок"""
    await message.answer(
        "📦 <b>Отчёт по поставкам</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("supplies"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("supplies:"))
async def supplies_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для поставок"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "📦 <b>Отчёт по поставкам</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "supplies", "platform": platform})


@dp.message_handler(commands=["feedbacks"])
async def feedbacks_command(message: types.Message):
    """Команда статистики отзывов"""
    await message.answer(
        "💬 <b>Статистика отзывов</b>\n\n" "Выберите платформу для анализа:",
        reply_markup=get_platform_keyboard("feedbacks"),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data.startswith("feedbacks:"))
async def feedbacks_platform_callback(callback: types.CallbackQuery):
    """Обработка выбора платформы для отзывов"""
    platform = callback.data.split(":")[1]

    await callback.message.edit_text(
        "💬 <b>Статистика отзывов</b>\n\n"
        f"Платформа: <b>{'Wildberries' if platform == 'wb' else 'Ozon' if platform == 'ozon' else 'Все'}</b>\n\n"
        "Выберите период:",
        reply_markup=get_period_keyboard(),
        parse_mode="HTML",
    )

    user_data = dp.current_state(user=callback.from_user.id)
    await user_data.set_data({"command": "feedbacks", "platform": platform})


@dp.message_handler(commands=["balance"])
async def balance_command(message: types.Message):
    """Команда баланса кошелька"""
    await message.answer("💰 <b>Загружаю данные баланса...</b>", parse_mode="HTML")
    await show_balance_report(message, "all")


# === COGS/OPEX Management Commands ===
@dp.message_handler(commands=["set_cogs"])
async def cmd_set_cogs(message: types.Message):
    """Set COGS for a specific SKU: /set_cogs SKU amount"""
    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer(
                "📊 <b>Установка COGS</b>\n\nИспользование: /set_cogs SKU сумма\n\nПример: /set_cogs ABC123 150.50",
                parse_mode="HTML",
            )
            return

        sku = args[0]
        try:
            amount = float(args[1])
        except ValueError:
            await message.answer(
                "❌ Сумма должна быть числом\n\nПример: /set_cogs ABC123 150.50", parse_mode="HTML"
            )
            return

        await message.answer(
            f"✅ <b>COGS записан</b>\n\nSKU: {sku}\nСебестоимость: {amount} ₽\n\n⚠️ Функция загрузки в разработке",
            parse_mode="HTML",
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)[:200]}", parse_mode="HTML")


@dp.message_handler(commands=["set_opex"])
async def cmd_set_opex(message: types.Message):
    """Set OPEX for a specific date: /set_opex category amount [date]"""
    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer(
                "💼 <b>Установка OPEX</b>\n\nИспользование: /set_opex категория сумма [дата]\n\nПример: /set_opex Аренда 50000\nИли: /set_opex Реклама 25000 2025-09-15",
                parse_mode="HTML",
            )
            return

        category = args[0]
        try:
            amount = float(args[1])
        except ValueError:
            await message.answer(
                "❌ Сумма должна быть числом\n\nПример: /set_opex Аренда 50000", parse_mode="HTML"
            )
            return

        # Get date (today if not specified)
        from datetime import datetime

        if len(args) >= 3:
            try:
                expense_date = datetime.strptime(args[2], "%Y-%m-%d").strftime("%Y-%m-%d")
            except ValueError:
                await message.answer(
                    "❌ Дата должна быть в формате YYYY-MM-DD\n\nПример: /set_opex Аренда 50000 2025-09-15",
                    parse_mode="HTML",
                )
                return
        else:
            expense_date = datetime.now().strftime("%Y-%m-%d")

        await message.answer(
            f"✅ <b>OPEX записан</b>\n\nКатегория: {category}\nСумма: {amount} ₽\nДата: {expense_date}\n\n⚠️ Функция загрузки в разработке",
            parse_mode="HTML",
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)[:200]}", parse_mode="HTML")


@dp.message_handler(commands=["view_cogs"])
async def cmd_view_cogs(message: types.Message):
    """View current COGS data"""
    try:
        # Get COGS data from database via API endpoint
        j = await http_async.get_json("/status/live", {})
        counts = j.get("counts", {})
        cogs_count = counts.get("cogs", 0)

        if cogs_count == 0:
            await message.answer(
                "📊 <b>COGS данные</b>\n\n❌ COGS данные не найдены\n\nДля добавления используйте: /set_cogs SKU сумма",
                parse_mode="HTML",
            )
        else:
            await message.answer(
                f"📊 <b>COGS данные</b>\n\n✅ Загружено записей: {cogs_count}\n\nДля добавления новых используйте: /set_cogs SKU сумма",
                parse_mode="HTML",
            )

    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)[:200]}", parse_mode="HTML")


@dp.message_handler(commands=["view_opex"])
async def cmd_view_opex(message: types.Message):
    """View current OPEX data"""
    try:
        # Get OPEX data from database via API endpoint
        j = await http_async.get_json("/status/live", {})
        counts = j.get("counts", {})
        opex_count = counts.get("opex", 0)

        if opex_count == 0:
            await message.answer(
                "💼 <b>OPEX данные</b>\n\n❌ OPEX данные не найдены\n\nДля добавления используйте: /set_opex категория сумма",
                parse_mode="HTML",
            )
        else:
            await message.answer(
                f"💼 <b>OPEX данные</b>\n\n✅ Загружено записей: {opex_count}\n\nДля добавления новых используйте: /set_opex категория сумма",
                parse_mode="HTML",
            )

    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)[:200]}", parse_mode="HTML")


async def show_ads_report(message, platform, date_from, date_to):
    """Показать отчёт по рекламе"""
    try:
        ads_data = await http_async.get_json(
            "/live/ads/overview",
            params={"platform": platform, "date_from": date_from, "date_to": date_to},
        )

        if not ads_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных рекламы</b>\n\n" "API недоступен или нет данных",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        # Раздельная выдача: WB / Ozon / Итого
        if platform == "all":
            wb_data = ads_data.get("wb", {})
            ozon_data = ads_data.get("ozon", {})
            combined_data = ads_data.get("combined", {})

            text = "📊 <b>Статистика рекламы</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            # WB секция
            text += "🟣 <b>Wildberries</b>\n"
            text += f"💰 Расходы: {_format_value(wb_data.get('spend'), 'rub')}\n"
            text += f"👁 Показы: {_format_value(wb_data.get('views'))}\n"
            text += f"🖱 Клики: {_format_value(wb_data.get('clicks'))}\n"
            text += f"📈 CTR: {_format_value(wb_data.get('ctr'), '%')}\n"
            text += f"💸 CPC: {_format_value(wb_data.get('cpc'), 'rub')}\n"
            text += f"🛒 Заказы: {_format_value(wb_data.get('orders'))}\n\n"

            # Ozon секция
            text += "🔵 <b>Ozon</b>\n"
            text += f"💰 Расходы: {_format_value(ozon_data.get('spend'), 'rub')}\n"
            text += f"👁 Показы: {_format_value(ozon_data.get('views'))}\n"
            text += f"🖱 Клики: {_format_value(ozon_data.get('clicks'))}\n"
            text += f"📈 CTR: {_format_value(ozon_data.get('ctr'), '%')}\n"
            text += f"💸 CPC: {_format_value(ozon_data.get('cpc'), 'rub')}\n"
            text += f"🛒 Заказы: {_format_value(ozon_data.get('orders'))}\n\n"

            # Итого
            text += "📊 <b>ИТОГО</b>\n"
            text += f"💰 Общие расходы: {_format_value(combined_data.get('spend'), 'rub')}\n"
            text += f"👁 Всего показов: {_format_value(combined_data.get('views'))}\n"
            text += f"🖱 Всего кликов: {_format_value(combined_data.get('clicks'))}\n"
            text += f"🛒 Всего заказов: {_format_value(combined_data.get('orders'))}\n"

        else:
            # Одна платформа
            text = "📊 <b>Статистика рекламы</b>\n"
            text += f"{platform_emoji} <b>{platform_name}</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            text += f"💰 <b>Расходы:</b> {_format_value(ads_data.get('spend'), 'rub')}\n"
            text += f"👁 <b>Показы:</b> {_format_value(ads_data.get('views'))}\n"
            text += f"🖱 <b>Клики:</b> {_format_value(ads_data.get('clicks'))}\n"
            text += f"📈 <b>CTR:</b> {_format_value(ads_data.get('ctr'), '%')}\n"
            text += f"💸 <b>CPC:</b> {_format_value(ads_data.get('cpc'), 'rub')}\n"
            text += f"🛒 <b>Заказы:</b> {_format_value(ads_data.get('orders'))}\n"

        text += f"\n<i>Обновлено: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_ads:{platform}:{date_from}:{date_to}"
            )
        )
        await message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке рекламы</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка рекламы: {e}")


async def show_supplies_report(message, platform, date_from, date_to):
    """Показать отчёт по поставкам"""
    try:
        supplies_data = await http_async.get_json(
            "/live/supplies",
            params={"platform": platform, "date_from": date_from, "date_to": date_to},
        )

        if not supplies_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных поставок</b>\n\n" "API недоступен или нет данных",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        # Раздельная выдача: WB / Ozon / Итого
        if platform == "all":
            wb_data = supplies_data.get("wb", [])
            ozon_data = supplies_data.get("ozon", [])

            text = "📦 <b>Отчёт по поставкам</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            # WB секция
            if wb_data:
                text += "🟣 <b>Wildberries</b>\n"
                wb_total_items = sum(item.get("quantity", 0) for item in wb_data)
                wb_total_cost = sum(item.get("cost_amount", 0) for item in wb_data)
                text += f"📦 Поставлено: {wb_total_items:,} шт.\n"
                text += f"💰 Общая стоимость: {await format_currency(wb_total_cost)}\n\n"
            else:
                text += "🟣 <b>Wildberries</b>\n📊 Нет данных\n\n"

            # Ozon секция
            if ozon_data:
                text += "🔵 <b>Ozon</b>\n"
                ozon_total_items = sum(item.get("quantity", 0) for item in ozon_data)
                ozon_total_cost = sum(item.get("cost_amount", 0) for item in ozon_data)
                text += f"📦 Поставлено: {ozon_total_items:,} шт.\n"
                text += f"💰 Общая стоимость: {await format_currency(ozon_total_cost)}\n\n"
            else:
                text += "🔵 <b>Ozon</b>\n📊 Нет данных\n\n"

            # Итого
            total_items = sum(item.get("quantity", 0) for item in wb_data + ozon_data)
            total_cost = sum(item.get("cost_amount", 0) for item in wb_data + ozon_data)
            text += "📊 <b>ИТОГО</b>\n"
            text += f"📦 Всего поставлено: {total_items:,} шт.\n"
            text += f"💰 Общая стоимость: {await format_currency(total_cost)}"

        else:
            # Одна платформа
            supplies = (
                supplies_data if isinstance(supplies_data, list) else supplies_data.get("data", [])
            )
            total_items = sum(item.get("quantity", 0) for item in supplies)
            total_cost = sum(item.get("cost_amount", 0) for item in supplies)

            text = "📦 <b>Отчёт по поставкам</b>\n"
            text += f"{platform_emoji} <b>{platform_name}</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            text += f"📦 <b>Поставлено:</b> {total_items:,} шт.\n"
            text += f"💰 <b>Общая стоимость:</b> {await format_currency(total_cost)}\n"

            if supplies:
                text += "\n🔝 <b>Топ поставки:</b>\n"
                top_supplies = sorted(supplies, key=lambda x: x.get("quantity", 0), reverse=True)[
                    :5
                ]

                for i, item in enumerate(top_supplies, 1):
                    name = item.get("name", item.get("vendor_code", "Без названия"))[:25]
                    quantity = item.get("quantity", 0)
                    cost = item.get("cost_amount", 0)

                    text += f"{i}. <b>{name}</b>\n"
                    text += f"   📦 {quantity} шт. × {await format_currency(cost)}\n"

        text += f"\n<i>Обновлено: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_supplies:{platform}:{date_from}:{date_to}"
            )
        )
        await message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке поставок</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка поставок: {e}")


async def show_feedbacks_report(message, platform, date_from, date_to):
    """Показать отчёт по отзывам"""
    try:
        feedbacks_data = await http_async.get_json(
            "/live/feedbacks",
            params={"platform": platform, "date_from": date_from, "date_to": date_to},
        )

        if not feedbacks_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки отзывов</b>\n\n" "API недоступен или нет данных",
                parse_mode="HTML",
            )
            return

        platform_emoji = {"wb": "🟣", "ozon": "🔵", "all": "📊"}[platform]
        platform_name = {"wb": "Wildberries", "ozon": "Ozon", "all": "Все платформы"}[platform]

        # Раздельная выдача: WB / Ozon / Итого
        if platform == "all":
            wb_data = feedbacks_data.get("wb", [])
            ozon_data = feedbacks_data.get("ozon", [])

            text = "💬 <b>Статистика отзывов</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            # WB секция
            if wb_data:
                text += "🟣 <b>Wildberries</b>\n"
                wb_avg_rating = sum(r.get("rating", 0) for r in wb_data) / len(wb_data)
                wb_answered = sum(1 for r in wb_data if r.get("is_answered"))
                text += f"⭐ Отзывов: {len(wb_data)}\n"
                text += f"📊 Средний рейтинг: {wb_avg_rating:.1f}\n"
                text += f"💬 Отвечено: {wb_answered}/{len(wb_data)}\n\n"
            else:
                text += "🟣 <b>Wildberries</b>\n📊 Нет данных\n\n"

            # Ozon секция
            if ozon_data:
                text += "🔵 <b>Ozon</b>\n"
                ozon_avg_rating = sum(r.get("rating", 0) for r in ozon_data) / len(ozon_data)
                ozon_answered = sum(1 for r in ozon_data if r.get("is_answered"))
                text += f"⭐ Отзывов: {len(ozon_data)}\n"
                text += f"📊 Средний рейтинг: {ozon_avg_rating:.1f}\n"
                text += f"💬 Отвечено: {ozon_answered}/{len(ozon_data)}\n\n"
            else:
                text += "🔵 <b>Ozon</b>\n📊 Нет данных\n\n"

            # Итого
            all_feedbacks = wb_data + ozon_data
            if all_feedbacks:
                total_avg_rating = sum(r.get("rating", 0) for r in all_feedbacks) / len(
                    all_feedbacks
                )
                total_answered = sum(1 for r in all_feedbacks if r.get("is_answered"))
                text += "📊 <b>ИТОГО</b>\n"
                text += f"⭐ Всего отзывов: {len(all_feedbacks)}\n"
                text += f"📊 Общий рейтинг: {total_avg_rating:.1f}\n"
                text += f"💬 Всего отвечено: {total_answered}/{len(all_feedbacks)}"

        else:
            # Одна платформа
            feedbacks = (
                feedbacks_data
                if isinstance(feedbacks_data, list)
                else feedbacks_data.get("data", [])
            )

            text = "💬 <b>Статистика отзывов</b>\n"
            text += f"{platform_emoji} <b>{platform_name}</b>\n"
            text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

            if feedbacks:
                avg_rating = sum(r.get("rating", 0) for r in feedbacks) / len(feedbacks)
                answered = sum(1 for r in feedbacks if r.get("is_answered"))

                text += f"⭐ <b>Отзывов:</b> {len(feedbacks)}\n"
                text += f"📊 <b>Средний рейтинг:</b> {avg_rating:.1f}\n"
                text += f"💬 <b>Отвечено:</b> {answered}/{len(feedbacks)}\n"

                # Последние отзывы
                text += "\n🕒 <b>Последние отзывы:</b>\n"
                recent_feedbacks = sorted(
                    feedbacks, key=lambda x: x.get("review_date", ""), reverse=True
                )[:3]

                for i, review in enumerate(recent_feedbacks, 1):
                    rating_stars = "⭐" * review.get("rating", 0)
                    review_text = review.get("text", "Без текста")[:50] + "..."
                    text += f"{i}. {rating_stars} {review_text}\n"
            else:
                text += "📊 <b>Нет отзывов за период</b>"

        text += f"\n<i>Обновлено: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"refresh_feedbacks:{platform}:{date_from}:{date_to}"
            )
        )
        await message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке отзывов</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка отзывов: {e}")


async def show_balance_report(message, platform):
    """Показать отчёт по балансу"""
    try:
        balance_data = await http_async.get_json("/live/balance", params={"platform": platform})

        if not balance_data:
            await message.edit_text(
                "❌ <b>Ошибка загрузки данных баланса</b>\n\n" "API недоступен или нет данных",
                parse_mode="HTML",
            )
            return

        # Раздельная выдача: WB / Ozon / Итого
        if platform == "all":
            wb_data = balance_data.get("wb", {})
            ozon_data = balance_data.get("ozon", {})

            text = "💰 <b>Баланс кошельков</b>\n"
            text += "🕐 <b>На данный момент</b>\n\n"

            # WB секция
            text += "🟣 <b>Wildberries</b>\n"
            if wb_data and not wb_data.get("missing_components"):
                text += f"💰 Баланс: {await format_currency(wb_data.get('balance', 0))}\n"
                text += f"🔒 К выводу: {await format_currency(wb_data.get('available', 0))}\n"
                text += f"⏳ В процессе: {await format_currency(wb_data.get('pending', 0))}\n\n"
            else:
                text += "📊 —\n\n"

            # Ozon секция
            text += "🔵 <b>Ozon</b>\n"
            if ozon_data and not ozon_data.get("missing_components"):
                text += f"💰 Баланс: {await format_currency(ozon_data.get('balance', 0))}\n"
                text += f"🔒 К выводу: {await format_currency(ozon_data.get('available', 0))}\n"
                text += f"⏳ В процессе: {await format_currency(ozon_data.get('pending', 0))}\n\n"
            else:
                text += "📊 —\n\n"

            # Итого (только если есть данные)
            has_wb_data = wb_data and not wb_data.get("missing_components")
            has_ozon_data = ozon_data and not ozon_data.get("missing_components")

            if has_wb_data or has_ozon_data:
                total_balance = (wb_data.get("balance", 0) if has_wb_data else 0) + (
                    ozon_data.get("balance", 0) if has_ozon_data else 0
                )
                total_available = (wb_data.get("available", 0) if has_wb_data else 0) + (
                    ozon_data.get("available", 0) if has_ozon_data else 0
                )

                text += "📊 <b>ИТОГО</b>\n"
                text += f"💰 Общий баланс: {await format_currency(total_balance)}\n"
                text += f"🔒 Всего к выводу: {await format_currency(total_available)}"
            else:
                text += "📊 <b>ИТОГО</b>\n📊 Нет доступных данных"

        else:
            # Одна платформа
            platform_emoji = {"wb": "🟣", "ozon": "🔵"}[platform]
            platform_name = {"wb": "Wildberries", "ozon": "Ozon"}[platform]

            text = "💰 <b>Баланс кошелька</b>\n"
            text += f"{platform_emoji} <b>{platform_name}</b>\n"
            text += "🕐 <b>На данный момент</b>\n\n"

            if balance_data and not balance_data.get("missing_components"):
                text += (
                    f"💰 <b>Баланс:</b> {await format_currency(balance_data.get('balance', 0))}\n"
                )
                text += f"🔒 <b>К выводу:</b> {await format_currency(balance_data.get('available', 0))}\n"
                text += f"⏳ <b>В процессе:</b> {await format_currency(balance_data.get('pending', 0))}\n"
            else:
                text += "📊 <b>Данные недоступны</b>"

        text += f"\n<i>Обновлено: {datetime.now().strftime('%H:%M')}</i>"

        kb = InlineKeyboardMarkup()
        kb.add(InlineKeyboardButton("🔄 Обновить", callback_data=f"refresh_balance:{platform}"))
        await message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.edit_text(
            f"❌ <b>Ошибка при загрузке баланса</b>\n\n" f"<code>{e!s}</code>", parse_mode="HTML"
        )
        logger.error(f"Ошибка баланса: {e}")


def _format_value(value, suffix=""):
    """Форматировать значение с поддержкой "—" для отсутствующих данных"""
    if value is None or (isinstance(value, (int, float)) and value == 0):
        return "—"

    if suffix == "rub":
        return f"{value:,.0f} ₽" if isinstance(value, (int, float)) else "—"
    elif suffix == "%":
        return f"{value:.2f}%" if isinstance(value, (int, float)) else "—"
    else:
        return f"{value:,}" if isinstance(value, (int, float)) else "—"


# Обработчики обновления отчётов
@dp.callback_query_handler(lambda c: c.data.startswith("refresh_"))
async def refresh_report_callback(callback: types.CallbackQuery):
    """Обновление отчётов"""
    parts = callback.data.split(":")
    report_type = parts[0].replace("refresh_", "")

    if report_type == "pnl" and len(parts) >= 4:
        platform, date_from, date_to = parts[1], parts[2], parts[3]
        from datetime import datetime

        date_from_obj = datetime.fromisoformat(date_from).date()
        date_to_obj = datetime.fromisoformat(date_to).date()

        # Определяем название периода
        if date_from == date_to:
            if date_from_obj == datetime.now().date():
                period_name = "Сегодня"
            else:
                period_name = "Вчера"
        else:
            period_name = f"{date_from} - {date_to}"

        await callback.message.edit_text("🔄 <b>Обновляю P&L...</b>", parse_mode="HTML")
        await show_pnl_report(callback.message, platform, date_from_obj, date_to_obj, period_name)

    elif report_type == "stock":
        platform = parts[1]
        await callback.message.edit_text("🔄 <b>Обновляю остатки...</b>", parse_mode="HTML")
        await show_stock_report(callback.message, platform)

    elif report_type == "ads" and len(parts) >= 4:
        platform, date_from, date_to = parts[1], parts[2], parts[3]
        await callback.message.edit_text("🔄 <b>Обновляю рекламу...</b>", parse_mode="HTML")
        await show_ads_report(callback.message, platform, date_from, date_to)

    elif report_type == "supplies" and len(parts) >= 4:
        platform, date_from, date_to = parts[1], parts[2], parts[3]
        await callback.message.edit_text("🔄 <b>Обновляю поставки...</b>", parse_mode="HTML")
        await show_supplies_report(callback.message, platform, date_from, date_to)

    elif report_type == "feedbacks" and len(parts) >= 4:
        platform, date_from, date_to = parts[1], parts[2], parts[3]
        await callback.message.edit_text("🔄 <b>Обновляю отзывы...</b>", parse_mode="HTML")
        await show_feedbacks_report(callback.message, platform, date_from, date_to)

    elif report_type == "balance":
        platform = parts[1]
        await callback.message.edit_text("🔄 <b>Обновляю баланс...</b>", parse_mode="HTML")
        await show_balance_report(callback.message, platform)

    # Аналогично для других типов отчётов...

    await callback.answer("✅ Отчёт обновлён")


# Система управления расходами
from expenses import CalculationType, ExpenseManager, ExpenseType, initialize_default_expenses

# Инициализация менеджера расходов
expense_manager = ExpenseManager()


@dp.message_handler(commands=["expenses"])
async def cmd_expenses_menu(message: types.Message):
    """Главное меню управления расходами"""
    summary = expense_manager.get_expense_summary()

    text = f"""💼 <b>Управление расходами</b>

📊 <b>Статистика:</b>
• Всего расходов: {summary['total_count']}
• Фиксированные расходы/месяц: {summary['monthly_fixed']:,.0f} ₽

📋 <b>По типам:</b>"""

    for expense_type, count in summary["by_type"].items():
        emoji = {
            "fixed": "📌",
            "commission": "💸",
            "logistics": "🚚",
            "penalty": "⚠️",
            "advertising": "📢",
            "other": "📝",
        }.get(expense_type, "📄")
        text += f"\n{emoji} {expense_type}: {count}"

    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("📝 Добавить расход", callback_data="expense_add"),
        InlineKeyboardButton("📋 Список расходов", callback_data="expense_list"),
    )
    kb.add(
        InlineKeyboardButton("📊 Расчет P&L", callback_data="expense_calculate"),
        InlineKeyboardButton("💰 Настроить COGS", callback_data="expense_cogs"),
    )
    kb.add(
        InlineKeyboardButton("📊 Шаблон COGS", callback_data="template_cogs"),
        InlineKeyboardButton("📊 Шаблон OPEX", callback_data="template_opex"),
    )
    kb.add(InlineKeyboardButton("⚙️ Настройки", callback_data="expense_settings"))

    await message.reply(text, parse_mode="HTML", reply_markup=kb)


@dp.message_handler(commands=["add_expense"])
async def cmd_add_expense(message: types.Message):
    """Команда добавления расхода"""
    args = message.get_args().strip()

    if not args:
        text = """📝 <b>Добавление расхода</b>

<b>Формат:</b>
<code>/add_expense [название] [тип] [способ_расчета] [сумма] [платформа] [категория]</code>

<b>Типы расходов:</b>
• fixed - постоянные
• commission - комиссии
• logistics - логистика
• penalty - штрафы
• advertising - реклама
• other - прочие

<b>Способы расчета:</b>
• fixed_amount - фиксированная сумма
• percent_revenue - % от выручки
• per_unit - за единицу товара
• per_order - за заказ

<b>Платформы:</b>
• wb, ozon, both, all (можно не указывать)

<b>Примеры:</b>
<code>/add_expense "Аренда офиса" fixed fixed_amount 25000</code>
<code>/add_expense "Комиссия WB" commission percent_revenue 15 wb</code>
<code>/add_expense "Упаковка" logistics per_unit 50 both packaging</code>"""

        await message.reply(text, parse_mode="HTML")
        return

    try:
        # Парсинг аргументов
        parts = []
        current = ""
        in_quotes = False

        for char in args:
            if char == '"' and not in_quotes:
                in_quotes = True
            elif char == '"' and in_quotes:
                in_quotes = False
                parts.append(current)
                current = ""
            elif char == " " and not in_quotes:
                if current:
                    parts.append(current)
                    current = ""
            else:
                current += char

        if current:
            parts.append(current)

        if len(parts) < 4:
            await message.reply(
                "❌ Недостаточно параметров. Используйте /add_expense без аргументов для помощи."
            )
            return

        name = parts[0]
        expense_type = ExpenseType(parts[1])
        calculation_type = CalculationType(parts[2])
        amount = float(parts[3])
        platform = (
            parts[4]
            if len(parts) > 4 and parts[4] not in ["all", "both"]
            else parts[4] if len(parts) > 4 else None
        )
        category = parts[5] if len(parts) > 5 else None

        expense_id = expense_manager.add_expense(
            name=name,
            expense_type=expense_type,
            calculation_type=calculation_type,
            amount=amount,
            platform=platform,
            category=category,
        )

        await message.reply(f"✅ Расход '{name}' добавлен с ID: {expense_id}")

    except ValueError as e:
        await message.reply(f"❌ Ошибка в параметрах: {e}")
    except Exception as e:
        await message.reply(f"❌ Ошибка: {e}")


@dp.message_handler(commands=["list_expenses"])
async def cmd_list_expenses(message: types.Message):
    """Список расходов"""
    args = message.get_args().strip()
    platform_filter = args if args in ["wb", "ozon", "both"] else None

    expenses = expense_manager.list_expenses(platform=platform_filter)

    if not expenses:
        await message.reply("📝 Расходы не найдены.")
        return

    text = "📋 <b>Список расходов</b>"
    if platform_filter:
        text += f" ({platform_filter})"
    text += "\n\n"

    for expense in expenses:
        emoji = {
            "fixed": "📌",
            "commission": "💸",
            "logistics": "🚚",
            "penalty": "⚠️",
            "advertising": "📢",
            "other": "📝",
        }.get(expense.expense_type.value, "📄")

        calc_text = {
            "fixed_amount": f"{expense.amount:,.0f} ₽",
            "percent_revenue": f"{expense.amount}%",
            "per_unit": f"{expense.amount:,.0f} ₽/шт",
            "per_order": f"{expense.amount:,.0f} ₽/заказ",
        }.get(expense.calculation_type.value, f"{expense.amount}")

        platform_text = f" [{expense.platform}]" if expense.platform else ""

        text += f"{emoji} <b>{expense.name}</b>{platform_text}\n"
        text += f"   {calc_text}"
        if expense.category:
            text += f" • {expense.category}"
        text += f"\n   <code>{expense.id}</code>\n\n"

    await message.reply(text, parse_mode="HTML")


@dp.message_handler(commands=["edit_expense"])
async def cmd_edit_expense(message: types.Message):
    """Редактирование расхода"""
    args = message.get_args().strip()

    if not args:
        await message.reply(
            "Использование: <code>/edit_expense [expense_id] [поле] [новое_значение]</code>",
            parse_mode="HTML",
        )
        return

    parts = args.split(maxsplit=2)
    if len(parts) < 3:
        await message.reply("❌ Укажите ID расхода, поле и новое значение")
        return

    expense_id, field, new_value = parts

    try:
        # Конвертируем значение в зависимости от поля
        if field == "amount":
            new_value = float(new_value)
        elif field == "expense_type":
            new_value = ExpenseType(new_value)
        elif field == "calculation_type":
            new_value = CalculationType(new_value)
        elif field == "is_active":
            new_value = new_value.lower() in ["true", "1", "да", "yes"]

        success = expense_manager.update_expense(expense_id, **{field: new_value})

        if success:
            await message.reply(f"✅ Расход {expense_id} обновлен")
        else:
            await message.reply(f"❌ Расход {expense_id} не найден")

    except Exception as e:
        await message.reply(f"❌ Ошибка: {e}")


@dp.message_handler(commands=["delete_expense"])
async def cmd_delete_expense(message: types.Message):
    """Удаление расхода"""
    expense_id = message.get_args().strip()

    if not expense_id:
        await message.reply(
            "Использование: <code>/delete_expense [expense_id]</code>", parse_mode="HTML"
        )
        return

    success = expense_manager.delete_expense(expense_id)

    if success:
        await message.reply(f"✅ Расход {expense_id} удален")
    else:
        await message.reply(f"❌ Расход {expense_id} не найден")


@dp.message_handler(commands=["init_expenses"])
async def cmd_init_expenses(message: types.Message):
    """Инициализация стандартных расходов"""
    if len(expense_manager.list_expenses()) > 0:
        await message.reply("⚠️ Расходы уже существуют. Используйте /expenses для управления.")
        return

    initialize_default_expenses(expense_manager)
    await message.reply(
        "✅ Стандартные расходы инициализированы. Используйте /expenses для просмотра."
    )


# Система анализа расходов и уведомлений
from expense_analyzer import ExpenseAnalyzer

# Инициализация анализатора расходов
expense_analyzer = ExpenseAnalyzer(expense_manager)


@dp.message_handler(commands=["check_penalties"])
async def cmd_check_penalties(message: types.Message):
    """Проверка штрафов за последние дни"""
    args = message.get_args().strip()
    days_back = int(args) if args.isdigit() else 7

    await message.reply("🔍 <b>Проверяю штрафы...</b>", parse_mode="HTML")

    try:
        alerts = await expense_analyzer.get_penalty_alerts(days_back=days_back)

        if not alerts:
            await message.reply(f"✅ Штрафы за последние {days_back} дней не обнаружены.")
            return

        text = f"🚨 <b>Обнаружено штрафов: {len(alerts)}</b>\n"
        text += f"📅 <b>За последние {days_back} дней</b>\n\n"

        total_penalties = sum(alert.amount for alert in alerts)
        text += f"💸 <b>Общая сумма штрафов:</b> {total_penalties:,.0f} ₽\n\n"

        severity_emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}

        for i, alert in enumerate(alerts[:10], 1):  # Показываем максимум 10
            emoji = severity_emoji.get(alert.severity, "⚠️")
            platform_emoji = {"ozon": "🔵", "wb": "🟣"}.get(alert.platform, "📱")

            text += f"{emoji} <b>{i}. {alert.reason}</b>\n"
            text += f"{platform_emoji} {alert.platform.upper()}: {alert.amount:,.0f} ₽\n"
            text += f"📅 {alert.date}\n"
            if alert.posting_number:
                text += f"📦 {alert.posting_number}\n"
            text += "\n"

        if len(alerts) > 10:
            text += f"... и еще {len(alerts) - 10} штрафов\n"

        text += f"<i>Данные обновлены: {datetime.now().strftime('%H:%M')}</i>"

        await message.reply(text, parse_mode="HTML")

    except Exception as e:
        await message.reply(f"❌ Ошибка при проверке штрафов: {e}")


@dp.message_handler(commands=["expense_report"])
async def cmd_expense_report(message: types.Message):
    """Отчет по расходам"""
    args = message.get_args().strip()
    days_back = int(args) if args.isdigit() else 30

    await message.reply("📊 <b>Генерирую отчет по расходам...</b>", parse_mode="HTML")

    try:
        from datetime import timedelta

        date_to = date.today()
        date_from = date_to - timedelta(days=days_back)

        report = await expense_analyzer.generate_expense_report(date_from, date_to)

        text = "📊 <b>Отчет по расходам</b>\n"
        text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

        text += f"💰 <b>Общие расходы:</b> {report['total_expenses']:,.0f} ₽\n\n"

        # По платформам
        text += "📱 <b>По платформам:</b>\n"
        for platform, data in report["platforms"].items():
            platform_emoji = {"ozon": "🔵", "wb": "🟣"}.get(platform, "📱")
            total = data.get("total_expenses", 0)
            count = data.get("expenses_count", 0)

            if "error" in data:
                text += f"{platform_emoji} {platform.upper()}: ⚠️ {data['error']}\n"
            else:
                text += f"{platform_emoji} {platform.upper()}: {total:,.0f} ₽ ({count} операций)\n"

        # Категории расходов
        categories = report["summary"].get("top_categories", {})
        if categories:
            text += "\n📋 <b>По категориям:</b>\n"
            category_emoji = {
                "commission": "💸",
                "logistics": "🚚",
                "advertising": "📢",
                "penalties": "⚠️",
                "returns": "↩️",
                "other": "📝",
            }

            for category, amount in list(categories.items())[:5]:
                emoji = category_emoji.get(category, "📄")
                text += f"{emoji} {category}: {amount:,.0f} ₽\n"

        # Штрафы
        penalties = report["summary"]
        if penalties["penalties_count"] > 0:
            text += f"\n🚨 <b>Штрафы:</b> {penalties['penalties_total']:,.0f} ₽ ({penalties['penalties_count']} шт.)\n"

        text += f"\n<i>Отчет сгенерирован: {datetime.now().strftime('%H:%M')}</i>"

        await message.reply(text, parse_mode="HTML")

    except Exception as e:
        await message.reply(f"❌ Ошибка генерации отчета: {e}")


@dp.message_handler(commands=["auto_penalties"])
async def cmd_auto_penalties(message: types.Message):
    """Включение/выключение автоматических уведомлений о штрафах"""
    # TODO: Реализовать систему автоматических уведомлений
    # Можно использовать планировщик задач или webhook
    await message.reply(
        "🔧 <b>Автоматические уведомления</b>\n\n"
        "Функция в разработке. Будет добавлена:\n"
        "• Ежедневная проверка штрафов\n"
        "• Уведомления при превышении лимитов\n"
        "• Настройка фильтров по сумме\n\n"
        "Используйте <code>/check_penalties</code> для ручной проверки.",
        parse_mode="HTML",
    )


# P&L калькулятор
from pnl_calculator import PnLCalculator

# Инициализация P&L калькулятора
pnl_calculator = PnLCalculator(expense_manager, expense_analyzer)


@dp.message_handler(commands=["pnl"])
async def cmd_pnl_report(message: types.Message):
    """P&L отчет с реальными расходами"""
    args = message.get_args().strip().split()

    # Параметры по умолчанию
    days_back = 30
    cost_per_unit = await get_user_default_cogs(
        message.from_user.id
    )  # Получаем сохраненную себестоимость

    # Парсинг аргументов
    if len(args) >= 1 and args[0].isdigit():
        days_back = int(args[0])
    if len(args) >= 2:
        try:
            cost_per_unit = float(args[1])
        except ValueError:
            pass

    await message.reply("📊 <b>Рассчитываю P&L отчет...</b>", parse_mode="HTML")

    try:
        from datetime import timedelta

        date_to = date.today()
        date_from = date_to - timedelta(days=days_back)

        # Рассчитываем P&L
        pnl = await pnl_calculator.calculate_pnl(date_from, date_to, cost_per_unit)

        # Форматируем отчет
        report_text = pnl_calculator.format_pnl_report(pnl, detailed=True)

        # Добавляем информацию о параметрах
        params_text = "\n📋 <b>Параметры расчета:</b>\n"
        params_text += f"• Период: {days_back} дней\n"
        params_text += f"• Себестоимость: {cost_per_unit:,.0f} ₽/шт\n"
        params_text += "• Использованы реальные данные из API\n\n"

        full_text = report_text + params_text

        # Создаем кнопки для дополнительных действий
        kb = InlineKeyboardMarkup(row_width=2)
        kb.add(
            InlineKeyboardButton(
                "💾 Сохранить отчет", callback_data=f"pnl_save:{days_back}:{cost_per_unit}"
            ),
            InlineKeyboardButton("📊 Детали расходов", callback_data=f"pnl_expenses:{days_back}"),
        )
        kb.add(
            InlineKeyboardButton(
                "🔄 Обновить", callback_data=f"pnl_refresh:{days_back}:{cost_per_unit}"
            )
        )

        await message.reply(full_text, parse_mode="HTML", reply_markup=kb)

    except Exception as e:
        await message.reply(f"❌ Ошибка расчета P&L: {e}")


@dp.message_handler(commands=["pnl_help"])
async def cmd_pnl_help(message: types.Message):
    """Справка по P&L отчетам"""
    text = """📊 <b>P&L отчеты - справка</b>

<b>Использование:</b>
<code>/pnl [дни] [себестоимость]</code>

<b>Параметры:</b>
• <b>дни</b> - период анализа (по умолчанию 30)
• <b>себестоимость</b> - COGS за единицу в рублях (использует сохраненное значение)

<b>Примеры:</b>
<code>/pnl</code> - отчет за 30 дней
<code>/pnl 7</code> - отчет за 7 дней
<code>/pnl 30 1200</code> - за 30 дней с COGS 1200₽

<b>Настройка себестоимости:</b>
<code>/set_default_cogs 1200</code> - установить COGS по умолчанию
<code>/view_default_cogs</code> - посмотреть текущую COGS

<b>Что включено в отчет:</b>
✅ Реальная выручка из API маркетплейсов
✅ Фактические расходы и комиссии
✅ Настроенные постоянные расходы
✅ Штрафы и пени
✅ Логистические расходы
✅ Рекламные расходы

<b>Другие команды:</b>
<code>/expenses</code> - управление расходами
<code>/expense_report</code> - детальный отчет по расходам
<code>/check_penalties</code> - проверка штрафов
<code>/set_cogs SKU сумма</code> - COGS для конкретного SKU"""

    await message.reply(text, parse_mode="HTML")


@dp.message_handler(commands=["set_default_cogs"])
async def cmd_set_default_cogs(message: types.Message):
    """Установка себестоимости по умолчанию для P&L расчетов"""
    args = message.get_args().strip()

    if not args:
        # Получаем текущую себестоимость
        current_cogs = await get_user_default_cogs(message.from_user.id)
        await message.reply(
            f"💰 <b>Установка себестоимости по умолчанию</b>\n\n"
            f"Использование: <code>/set_default_cogs [сумма]</code>\n"
            f"Пример: <code>/set_default_cogs 1200</code>\n\n"
            f"Текущая себестоимость: {current_cogs:,.0f} ₽/шт",
            parse_mode="HTML",
        )
        return

    try:
        cost = float(args)
        if cost <= 0:
            await message.reply("❌ Себестоимость должна быть больше 0")
            return

        # Сохраняем в настройки пользователя
        await save_user_default_cogs(message.from_user.id, cost)
        await message.reply(f"✅ Себестоимость по умолчанию установлена: {cost:,.0f} ₽/шт")

    except ValueError:
        await message.reply("❌ Неверный формат суммы. Укажите число.")


# Функции для работы с настройками пользователя
import json
import os


async def get_user_default_cogs(user_id: int) -> float:
    """Получение себестоимости по умолчанию для пользователя"""
    try:
        settings_file = f"data/user_settings_{user_id}.json"
        logger.info(f"Читаю настройки COGS из {settings_file}")
        if os.path.exists(settings_file):
            with open(settings_file, encoding="utf-8") as f:
                settings = json.load(f)
                cogs = float(settings.get("default_cogs", 800.0))
                logger.info(f"Загружена COGS из файла: {cogs}")
                return cogs
        else:
            logger.info(f"Файл настроек не найден: {settings_file}")
    except Exception as e:
        logger.error(f"Ошибка чтения настроек COGS: {e}")
    logger.info("Возвращаю COGS по умолчанию: 800.0")
    return 800.0  # Значение по умолчанию


async def save_user_default_cogs(user_id: int, cogs: float):
    """Сохранение себестоимости по умолчанию для пользователя"""
    try:
        os.makedirs("data", exist_ok=True)
        settings_file = f"data/user_settings_{user_id}.json"

        settings = {}
        if os.path.exists(settings_file):
            with open(settings_file, encoding="utf-8") as f:
                settings = json.load(f)

        settings["default_cogs"] = cogs
        settings["updated_at"] = datetime.now().isoformat()

        with open(settings_file, "w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)

    except Exception as e:
        logger.error(f"Ошибка сохранения настроек пользователя: {e}")


@dp.message_handler(commands=["view_default_cogs"])
async def cmd_view_default_cogs(message: types.Message):
    """Просмотр текущей себестоимости по умолчанию"""
    current_cogs = await get_user_default_cogs(message.from_user.id)

    text = "💰 <b>Себестоимость по умолчанию</b>\n\n"
    text += f"📊 Текущее значение: {current_cogs:,.0f} ₽/шт\n\n"
    text += "Используется в командах:\n"
    text += "• <code>/pnl</code> - для расчета P&L\n\n"
    text += "Для изменения: <code>/set_default_cogs [сумма]</code>"

    await message.reply(text, parse_mode="HTML")


async def generate_cogs_template_from_api(user_id: int) -> str | None:
    """Генерация шаблона COGS на основе SKU из API"""
    try:
        import os
        from datetime import datetime

        import pandas as pd

        logger.info(f"Генерируем шаблон COGS для пользователя {user_id}")

        # Собираем SKU из всех источников
        all_skus = set()

        # Получаем SKU из WB
        try:
            from api_clients_main import wb_business_api

            wb_cards = await wb_business_api.get_product_cards()
            for card in wb_cards:
                # Извлекаем SKU из карточки товара
                sizes = card.get("sizes", [])
                for size in sizes:
                    for sku_data in size.get("skus", []):
                        sku = sku_data.get("skus", [])
                        if isinstance(sku, list) and sku:
                            all_skus.add(str(sku[0]))
                        elif isinstance(sku, (str, int)):
                            all_skus.add(str(sku))

            logger.info(f"Получено {len(all_skus)} SKU из WB")
        except Exception as e:
            logger.error(f"Ошибка получения SKU из WB: {e}")

        # Получаем SKU из Ozon
        try:
            from api_clients_main import ozon_api

            ozon_stocks = await ozon_api.get_product_stocks()
            for item in ozon_stocks:  # Переименовываем stock в item для ясности
                offer_id = item.get("offer_id")  # Используем поле 'offer_id'
                if offer_id:
                    all_skus.add(str(offer_id))

            logger.info(f"Добавлено SKU из Ozon, всего: {len(all_skus)}")
        except Exception as e:
            logger.error(f"Ошибка получения SKU из Ozon: {e}")

        if not all_skus:
            logger.warning("Не найдено SKU из API, создаем демо-шаблон")
            # Создаем демо-шаблон если нет данных
            all_skus = {"182906671", "209962134", "182906670", "1483568647", "1483568688"}

        # Получаем текущую COGS по умолчанию для заполнения
        default_cogs = await get_user_default_cogs(user_id)

        # Создаем DataFrame
        sku_list = sorted(list(all_skus))
        df_data = []

        for sku in sku_list:
            df_data.append(
                {
                    "SKU": sku,
                    "COGS_RUB": default_cogs,  # Заполняем дефолтным значением
                    "Comment": "Укажите себестоимость товара",
                }
            )

        df = pd.DataFrame(df_data)

        # Сохраняем файл
        os.makedirs("templates", exist_ok=True)
        filename = (
            f"templates/cogs_template_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        # Создаем Excel файл с форматированием
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="COGS_Template", index=False)

            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets["COGS_Template"]

            # Устанавливаем ширину колонок
            worksheet.column_dimensions["A"].width = 15  # SKU
            worksheet.column_dimensions["B"].width = 12  # COGS_RUB
            worksheet.column_dimensions["C"].width = 35  # Comment

            # Добавляем заголовки с инструкциями
            worksheet.insert_rows(1, 3)
            worksheet["A1"] = "ШАБЛОН СЕБЕСТОИМОСТИ (COGS)"
            worksheet["A2"] = f'Сгенерирован: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            worksheet["A3"] = "Инструкция: Заполните колонку COGS_RUB и отправьте файл обратно боту"

            # Стилизация заголовков
            from openpyxl.styles import Alignment, Font, PatternFill

            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            worksheet["A1"].font = title_font
            worksheet["A2"].font = Font(size=10)
            worksheet["A3"].font = Font(size=10, italic=True)

            # Форматируем заголовки таблицы (строка 4)
            for col in ["A", "B", "C"]:
                cell = worksheet[f"{col}4"]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        logger.info(f"Шаблон COGS создан: {filename} ({len(sku_list)} SKU)")
        return filename

    except Exception as e:
        logger.error(f"Ошибка генерации шаблона COGS: {e}")
        return None


async def generate_opex_template() -> str | None:
    """Генерация шаблона OPEX с типовыми статьями расходов"""
    try:
        import os
        from datetime import datetime

        import pandas as pd

        logger.info("Генерируем шаблон OPEX")

        # Типовые статьи операционных расходов для маркетплейса
        opex_categories = [
            {
                "Category": "Маркетинг",
                "Subcategory": "Контекстная реклама WB",
                "Amount_RUB": 0,
                "Description": "Расходы на рекламу в WB",
            },
            {
                "Category": "Маркетинг",
                "Subcategory": "Контекстная реклама Ozon",
                "Amount_RUB": 0,
                "Description": "Расходы на рекламу в Ozon",
            },
            {
                "Category": "Маркетинг",
                "Subcategory": "Продвижение в соц. сетях",
                "Amount_RUB": 0,
                "Description": "SMM, таргетированная реклама",
            },
            {
                "Category": "Маркетинг",
                "Subcategory": "Блогеры и инфлюенсеры",
                "Amount_RUB": 0,
                "Description": "Сотрудничество с блогерами",
            },
            {
                "Category": "Логистика",
                "Subcategory": "Доставка на склад WB",
                "Amount_RUB": 0,
                "Description": "Стоимость доставки товаров на FBS",
            },
            {
                "Category": "Логистика",
                "Subcategory": "Доставка на склад Ozon",
                "Amount_RUB": 0,
                "Description": "Стоимость доставки товаров на FBS",
            },
            {
                "Category": "Логистика",
                "Subcategory": "Упаковочные материалы",
                "Amount_RUB": 0,
                "Description": "Коробки, пакеты, скотч",
            },
            {
                "Category": "Логистика",
                "Subcategory": "Хранение товаров",
                "Amount_RUB": 0,
                "Description": "Аренда склада/кладовки",
            },
            {
                "Category": "Операционные",
                "Subcategory": "Комиссия маркетплейсов",
                "Amount_RUB": 0,
                "Description": "Комиссии WB/Ozon (если не в отчетах)",
            },
            {
                "Category": "Операционные",
                "Subcategory": "Банковские услуги",
                "Amount_RUB": 0,
                "Description": "Обслуживание счета, эквайринг",
            },
            {
                "Category": "Операционные",
                "Subcategory": "Бухгалтерские услуги",
                "Amount_RUB": 0,
                "Description": "Ведение учета, отчетность",
            },
            {
                "Category": "Операционные",
                "Subcategory": "ИТ-сервисы",
                "Amount_RUB": 0,
                "Description": "CRM, аналитика, автоматизация",
            },
            {
                "Category": "Персонал",
                "Subcategory": "Зарплата сотрудников",
                "Amount_RUB": 0,
                "Description": "ФОТ основного персонала",
            },
            {
                "Category": "Персонал",
                "Subcategory": "Услуги фрилансеров",
                "Amount_RUB": 0,
                "Description": "Дизайн, копирайтинг, фото",
            },
            {
                "Category": "Персонал",
                "Subcategory": "Обучение персонала",
                "Amount_RUB": 0,
                "Description": "Курсы, семинары, тренинги",
            },
            {
                "Category": "Прочие",
                "Subcategory": "Офисные расходы",
                "Amount_RUB": 0,
                "Description": "Аренда, коммунальные, интернет",
            },
            {
                "Category": "Прочие",
                "Subcategory": "Юридические услуги",
                "Amount_RUB": 0,
                "Description": "Консультации, договоры",
            },
            {
                "Category": "Прочие",
                "Subcategory": "Прочие операционные",
                "Amount_RUB": 0,
                "Description": "Непредвиденные расходы",
            },
        ]

        df = pd.DataFrame(opex_categories)

        # Сохраняем файл
        os.makedirs("templates", exist_ok=True)
        filename = f"templates/opex_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Создаем Excel файл с форматированием
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="OPEX_Template", index=False)

            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets["OPEX_Template"]

            # Устанавливаем ширину колонок
            worksheet.column_dimensions["A"].width = 15  # Category
            worksheet.column_dimensions["B"].width = 25  # Subcategory
            worksheet.column_dimensions["C"].width = 12  # Amount_RUB
            worksheet.column_dimensions["D"].width = 40  # Description

            # Добавляем заголовки с инструкциями
            worksheet.insert_rows(1, 3)
            worksheet["A1"] = "ШАБЛОН ОПЕРАЦИОННЫХ РАСХОДОВ (OPEX)"
            worksheet["A2"] = f'Сгенерирован: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            worksheet["A3"] = "Инструкция: Заполните колонку Amount_RUB суммами ваших расходов"

            # Стилизация заголовков
            from openpyxl.styles import Alignment, Font, PatternFill

            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            worksheet["A1"].font = title_font
            worksheet["A2"].font = Font(size=10)
            worksheet["A3"].font = Font(size=10, italic=True)

            # Форматируем заголовки таблицы (строка 4)
            for col in ["A", "B", "C", "D"]:
                cell = worksheet[f"{col}4"]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        logger.info(f"Шаблон OPEX создан: {filename} ({len(opex_categories)} категорий)")
        return filename

    except Exception as e:
        logger.error(f"Ошибка генерации шаблона OPEX: {e}")
        return None


# Обработчик callback-ов для P&L
@dp.callback_query_handler(lambda c: c.data.startswith("pnl_"))
async def process_pnl_callback(callback: types.CallbackQuery):
    """Обработка callback-ов P&L отчетов"""
    data_parts = callback.data.split(":")
    action = data_parts[0]

    if action == "pnl_save":
        days_back = int(data_parts[1])
        cost_per_unit = float(data_parts[2])

        try:
            from datetime import timedelta

            date_to = date.today()
            date_from = date_to - timedelta(days=days_back)

            pnl = await pnl_calculator.calculate_pnl(date_from, date_to, cost_per_unit)
            filename = await pnl_calculator.save_pnl_report(pnl)

            await callback.message.reply(f"💾 P&L отчет сохранен: {filename}")

        except Exception as e:
            await callback.message.reply(f"❌ Ошибка сохранения: {e}")

    elif action == "pnl_expenses":
        days_back = int(data_parts[1])
        await callback.message.edit_text("🔄 <b>Загружаю детали расходов...</b>", parse_mode="HTML")

        # Показываем детальный отчет по расходам
        from datetime import timedelta

        date_to = date.today()
        date_from = date_to - timedelta(days=days_back)

        report = await expense_analyzer.generate_expense_report(date_from, date_to)

        text = "📊 <b>Детали расходов</b>\n"
        text += f"📅 <b>{date_from} — {date_to}</b>\n\n"

        for platform, data in report["platforms"].items():
            platform_emoji = {"ozon": "🔵", "wb": "🟣"}.get(platform, "📱")
            total = data.get("total_expenses", 0)

            if total > 0:
                text += f"{platform_emoji} <b>{platform.upper()}: {total:,.0f} ₽</b>\n"
                categories = data.get("categories", {})
                for category, amount in categories.items():
                    if amount > 0:
                        text += f"  • {category}: {amount:,.0f} ₽\n"
                text += "\n"

        await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "pnl_refresh":
        days_back = int(data_parts[1])
        cost_per_unit = float(data_parts[2])

        await callback.message.edit_text("🔄 <b>Обновляю P&L отчет...</b>", parse_mode="HTML")

        # Повторно вызываем расчет P&L
        # Аналогично основной команде
        pass

    await callback.answer("✅ Готово")


# Обработчик callback-ов для управления расходами
@dp.callback_query_handler(lambda c: c.data.startswith("expense_"))
async def process_expense_callback(callback: types.CallbackQuery):
    """Обработка callback-ов управления расходами"""
    action = callback.data
    logger.info(f"Получен expense callback: {action} от пользователя {callback.from_user.id}")

    await callback.answer()  # Убираем индикатор "загрузка"

    if action == "expense_add":
        text = """📝 <b>Добавление расхода</b>

<b>Формат команды:</b>
<code>/add_expense "[название]" [тип] [способ_расчета] [сумма] [платформа] [категория]</code>

<b>Типы расходов:</b>
• fixed - постоянные
• commission - комиссии
• logistics - логистика
• penalty - штрафы
• advertising - реклама
• other - прочие

<b>Способы расчета:</b>
• fixed_amount - фиксированная сумма
• percent_revenue - % от выручки
• per_unit - за единицу товара
• per_order - за заказ

<b>Примеры:</b>
<code>/add_expense "Аренда офиса" fixed fixed_amount 25000</code>
<code>/add_expense "Комиссия WB" commission percent_revenue 27 wb</code>"""

        await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "expense_list":
        await callback.message.edit_text("🔄 <b>Загружаю список расходов...</b>", parse_mode="HTML")

        expenses = expense_manager.list_expenses()

        if not expenses:
            text = "📝 <b>Список расходов</b>\n\n❌ Расходы не найдены.\n\nИспользуйте /add_expense для добавления."
        else:
            text = "📋 <b>Список расходов</b>\n\n"

            for expense in expenses[:10]:  # Показываем первые 10
                emoji = {
                    "fixed": "📌",
                    "commission": "💸",
                    "logistics": "🚚",
                    "penalty": "⚠️",
                    "advertising": "📢",
                    "other": "📝",
                }.get(expense.expense_type.value, "📄")

                calc_text = {
                    "fixed_amount": f"{expense.amount:,.0f} ₽",
                    "percent_revenue": f"{expense.amount}%",
                    "per_unit": f"{expense.amount:,.0f} ₽/шт",
                    "per_order": f"{expense.amount:,.0f} ₽/заказ",
                }.get(expense.calculation_type.value, f"{expense.amount}")

                platform_text = f" [{expense.platform}]" if expense.platform else ""

                text += f"{emoji} <b>{expense.name}</b>{platform_text}\n"
                text += f"   {calc_text}\n\n"

            if len(expenses) > 10:
                text += f"... и еще {len(expenses) - 10} расходов\n\n"

            text += "Для полного списка: <code>/list_expenses</code>"

        await callback.message.edit_text(text, parse_mode="HTML")

    elif action == "expense_calculate":
        await callback.message.edit_text(
            "📊 <b>Для расчета P&L используйте:</b>\n\n<code>/pnl</code> - полный P&L отчет\n<code>/pnl_help</code> - справка",
            parse_mode="HTML",
        )

    elif action == "expense_cogs":
        try:
            logger.info(f"Обработка expense_cogs для пользователя {callback.from_user.id}")
            current_cogs = await get_user_default_cogs(callback.from_user.id)
            logger.info(f"Текущая COGS для пользователя {callback.from_user.id}: {current_cogs}")

            text = "💰 <b>Настройка себестоимости</b>\n\n"
            text += f"📊 Текущее значение: {current_cogs:,.0f} ₽/шт\n\n"
            text += "<b>Для изменения используйте:</b>\n"
            text += "<code>/set_default_cogs [сумма]</code>\n\n"
            text += "<b>Примеры:</b>\n"
            text += "<code>/set_default_cogs 1200</code>\n"
            text += "<code>/set_default_cogs 800</code>\n\n"
            text += "<i>Себестоимость используется в P&L расчетах</i>"

            await callback.message.edit_text(text, parse_mode="HTML")
            logger.info(f"COGS сообщение отправлено пользователю {callback.from_user.id}")
        except Exception as e:
            logger.error(f"Ошибка в обработке expense_cogs: {e}")
            await callback.message.edit_text(
                "❌ Произошла ошибка при загрузке настроек", parse_mode="HTML"
            )

    elif action == "expense_settings":
        summary = expense_manager.get_expense_summary()

        text = "⚙️ <b>Настройки расходов</b>\n\n"
        text += "📊 <b>Статистика:</b>\n"
        text += f"• Всего расходов: {summary['total_count']}\n"
        text += f"• Фиксированные/месяц: {summary['monthly_fixed']:,.0f} ₽\n\n"
        text += "<b>Команды управления:</b>\n"
        text += "<code>/expenses</code> - главное меню\n"
        text += "<code>/init_expenses</code> - добавить стандартные расходы\n"
        text += "<code>/expense_report</code> - отчет по расходам\n"
        text += "<code>/check_penalties</code> - проверить штрафы"

        await callback.message.edit_text(text, parse_mode="HTML")

    await callback.answer("✅ Готово")


# Обработчик callback-ов для генерации шаблонов
@dp.callback_query_handler(lambda c: c.data.startswith("template_"))
async def process_template_callback(callback: types.CallbackQuery):
    """Обработка callback-ов генерации шаблонов"""
    action = callback.data
    logger.info(f"Получен template callback: {action} от пользователя {callback.from_user.id}")

    await callback.answer()  # Убираем индикатор "загрузка"

    if action == "template_cogs":
        await callback.message.edit_text(
            "📊 <b>Генерирую шаблон COGS...</b>\n\nПолучаю SKU из API маркетплейсов...",
            parse_mode="HTML",
        )

        try:
            filename = await generate_cogs_template_from_api(callback.from_user.id)

            if filename:
                # Отправляем файл пользователю
                with open(filename, "rb") as doc:
                    await callback.message.reply_document(
                        doc,
                        caption="📊 <b>Шаблон COGS сгенерирован!</b>\n\n"
                        "📋 В файле содержатся все ваши SKU из API WB/Ozon\n"
                        "💰 Заполнено текущими значениями COGS по умолчанию\n"
                        "📝 Измените суммы и отправьте файл обратно боту\n\n"
                        "<b>Инструкция:</b>\n"
                        "1. Откройте файл в Excel/Calc\n"
                        "2. Заполните колонку COGS_RUB\n"
                        "3. Сохраните файл\n"
                        "4. Отправьте его в бот для загрузки",
                        parse_mode="HTML",
                    )

                # Удаляем временный файл
                import os

                try:
                    os.remove(filename)
                except:
                    pass
            else:
                await callback.message.edit_text(
                    "❌ <b>Ошибка генерации шаблона COGS</b>\n\n"
                    "Попробуйте позже или обратитесь к администратору",
                    parse_mode="HTML",
                )
        except Exception as e:
            logger.error(f"Ошибка генерации шаблона COGS: {e}")
            await callback.message.edit_text(
                f"❌ <b>Ошибка генерации шаблона</b>\n\n{str(e)[:100]}", parse_mode="HTML"
            )

    elif action == "template_opex":
        await callback.message.edit_text(
            "📊 <b>Генерирую шаблон OPEX...</b>\n\nСоздаю таблицу с типовыми статьями расходов...",
            parse_mode="HTML",
        )

        try:
            filename = await generate_opex_template()

            if filename:
                # Отправляем файл пользователю
                with open(filename, "rb") as doc:
                    await callback.message.reply_document(
                        doc,
                        caption="📊 <b>Шаблон OPEX сгенерирован!</b>\n\n"
                        "📋 В файле содержатся типовые статьи операционных расходов\n"
                        "💼 Разделены по категориям: маркетинг, логистика, персонал, прочие\n"
                        "📝 Заполните суммы и отправьте файл обратно боту\n\n"
                        "<b>Инструкция:</b>\n"
                        "1. Откройте файл в Excel/Calc\n"
                        "2. Заполните колонку Amount_RUB\n"
                        "3. Добавьте/удалите строки по необходимости\n"
                        "4. Сохраните файл\n"
                        "5. Отправьте его в бот для загрузки",
                        parse_mode="HTML",
                    )

                # Удаляем временный файл
                import os

                try:
                    os.remove(filename)
                except:
                    pass
            else:
                await callback.message.edit_text(
                    "❌ <b>Ошибка генерации шаблона OPEX</b>\n\n"
                    "Попробуйте позже или обратитесь к администратору",
                    parse_mode="HTML",
                )
        except Exception as e:
            logger.error(f"Ошибка генерации шаблона OPEX: {e}")
            await callback.message.edit_text(
                f"❌ <b>Ошибка генерации шаблона</b>\n\n{str(e)[:100]}", parse_mode="HTML"
            )


# ============= НОВЫЕ ОБРАБОТЧИКИ EXCEL ЭКСПОРТА =============


@dp.callback_query_handler(lambda c: c.data.startswith("export_dds_excel:"))
async def excel_dds_export_callback(callback_query: types.CallbackQuery):
    """Обработчик экспорта DDS в Excel"""
    await handle_dds_excel_export(callback_query)


@dp.callback_query_handler(lambda c: c.data.startswith("export_pnl_excel:"))
async def excel_pnl_export_callback(callback_query: types.CallbackQuery):
    """Обработчик экспорта P&L в Excel"""
    await handle_pnl_excel_export(callback_query)


@dp.callback_query_handler(lambda c: c.data == "upload_cost_template")
async def cost_template_callback(callback_query: types.CallbackQuery):
    """Обработчик запроса шаблона себестоимости"""
    await handle_cost_template_upload(callback_query)


@dp.message_handler(content_types=[types.ContentType.DOCUMENT])
async def document_handler(message: types.Message):
    """Обработчик загруженных документов (включая файлы себестоимости)"""
    # Проверяем, является ли это файлом себестоимости
    if message.document and message.document.file_name:
        filename = message.document.file_name.lower()
        if filename.endswith((".xlsx", ".xls")) and (
            "cost" in filename or "себестоимость" in filename.lower() or "template" in filename
        ):
            await handle_cost_file_upload(message)
            return

    # Если это не файл себестоимости, обрабатываем как обычный документ
    await message.reply(
        "📄 <b>Документ получен</b>\n\n"
        "Если это файл себестоимости, убедитесь что имя файла содержит слово 'cost', 'template' или 'себестоимость'",
        parse_mode="HTML",
    )


@dp.message_handler(text="💰 Сводка себестоимости")
async def cost_summary_handler(message: types.Message):
    """Показать сводку по загруженным данным о себестоимости"""
    try:
        summary = await generate_cost_summary_for_bot()
        await message.answer(summary, parse_mode="HTML")
    except Exception as e:
        logger.error(f"Ошибка генерации сводки себестоимости: {e}")
        await message.answer(
            "❌ <b>Ошибка получения сводки себестоимости</b>\n\n"
            "Возможно, данные о себестоимости еще не загружены.",
            parse_mode="HTML",
        )


# === НОВЫЕ CALLBACK ОБРАБОТЧИКИ ===


@dp.callback_query_handler(lambda c: c.data == "main_menu")
async def main_menu_callback(callback_query: types.CallbackQuery):
    """Возврат в главное меню"""
    await callback_query.answer()
    welcome_text = (
        "🏠 <b>Главное меню</b>\n\n"
        "🟣 <b>Отчеты WB</b> - финансовые отчеты Wildberries\n"
        "🟠 <b>Отчеты Ozon</b> - финансовые отчеты Ozon\n"
        "📋 <b>Загрузка данных WB</b> - загрузка Excel файлов для анализа\n"
        "💰 <b>Себестоимость</b> - управление данными о себестоимости\n"
        "⭐ <b>Управление отзывами</b> - автоответы на отзывы\n"
        "🔍 <b>API статус</b> - проверка подключений к маркетплейсам"
    )
    await callback_query.message.answer(
        welcome_text, parse_mode="HTML", reply_markup=get_main_menu()
    )


# WB Отчеты
@dp.callback_query_handler(lambda c: c.data == "wb_financial")
async def wb_financial_callback(callback_query: types.CallbackQuery):
    """WB Финансовый отчет"""
    await callback_query.answer()
    from date_picker import date_range_manager, get_enhanced_period_menu

    date_range_manager.start_date_selection(callback_query.from_user.id, "wb_financial")

    await callback_query.message.edit_text(
        "🟣 <b>Wildberries - Финансовый отчет</b>\n\n"
        "📅 Выберите период для анализа:\n"
        "⚠️ <i>Максимальный период: 176 дней (WB)</i>",
        reply_markup=get_enhanced_period_menu(),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data == "wb_cumulative")
async def wb_cumulative_callback(callback_query: types.CallbackQuery):
    """WB Нарастающий итог"""
    await callback_query.answer()
    from date_picker import date_range_manager, get_enhanced_period_menu

    date_range_manager.start_date_selection(callback_query.from_user.id, "wb_cumulative")

    await callback_query.message.edit_text(
        "🟣 <b>Wildberries - Нарастающий итог</b>\n\n"
        "📅 Выберите период для анализа:\n"
        "⚠️ <i>Максимальный период: 176 дней (WB)</i>",
        reply_markup=get_enhanced_period_menu(),
        parse_mode="HTML",
    )


# Ozon Отчеты
@dp.callback_query_handler(lambda c: c.data == "ozon_financial")
async def ozon_financial_callback(callback_query: types.CallbackQuery):
    """Ozon Финансовый отчет"""
    await callback_query.answer()
    from date_picker import date_range_manager, get_enhanced_period_menu

    date_range_manager.start_date_selection(callback_query.from_user.id, "ozon_financial")

    await callback_query.message.edit_text(
        "🟠 <b>Ozon - Финансовый отчет</b>\n\n"
        "📅 Выберите период для анализа:\n"
        "✅ <i>Доступны периоды до 180 дней</i>",
        reply_markup=get_enhanced_period_menu(),
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data == "ozon_cumulative")
async def ozon_cumulative_callback(callback_query: types.CallbackQuery):
    """Ozon Нарастающий итог"""
    await callback_query.answer()
    from date_picker import date_range_manager, get_enhanced_period_menu

    date_range_manager.start_date_selection(callback_query.from_user.id, "ozon_cumulative")

    await callback_query.message.edit_text(
        "🟠 <b>Ozon - Нарастающий итог</b>\n\n"
        "📅 Выберите период для анализа:\n"
        "✅ <i>Доступны периоды до 180 дней</i>",
        reply_markup=get_enhanced_period_menu(),
        parse_mode="HTML",
    )


# WB Загрузка файлов
@dp.callback_query_handler(lambda c: c.data == "wb_upload_sales")
async def wb_upload_sales_callback(callback_query: types.CallbackQuery):
    """Загрузка отчета о продажах WB"""
    await callback_query.answer()

    # Устанавливаем состояние ожидания файла
    user_waiting_for_file[callback_query.from_user.id] = "sales"

    await callback_query.message.edit_text(
        "📤 <b>Загрузка отчета о продажах WB</b>\n\n"
        "📋 Отправьте Excel файл с отчетом о продажах, скачанный из личного кабинета Wildberries.\n\n"
        "💡 <b>Как получить файл:</b>\n"
        "1. Зайдите в личный кабинет WB\n"
        '2. Перейдите в "Аналитика" → "Отчеты"\n'
        '3. Выберите "Отчет о продажах по периодам"\n'
        "4. Скачайте Excel файл\n"
        "5. Отправьте его сюда\n\n"
        "⏳ <b>Жду файл...</b>",
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data == "wb_upload_orders")
async def wb_upload_orders_callback(callback_query: types.CallbackQuery):
    """Загрузка отчета о заказах WB"""
    await callback_query.answer()

    # Устанавливаем состояние ожидания файла
    user_waiting_for_file[callback_query.from_user.id] = "orders"

    await callback_query.message.edit_text(
        "📤 <b>Загрузка отчета о заказах WB</b>\n\n"
        "📋 Отправьте Excel файл с отчетом о заказах, скачанный из личного кабинета Wildberries.\n\n"
        "💡 <b>Как получить файл:</b>\n"
        "1. Зайдите в личный кабинет WB\n"
        '2. Перейдите в "Аналитика" → "Отчеты"\n'
        '3. Выберите "Отчет о заказах"\n'
        "4. Скачайте Excel файл\n"
        "5. Отправьте его сюда\n\n"
        "⏳ <b>Жду файл...</b>",
        parse_mode="HTML",
    )


@dp.callback_query_handler(lambda c: c.data == "wb_upload_finance")
async def wb_upload_finance_callback(callback_query: types.CallbackQuery):
    """Загрузка финансового отчета WB"""
    await callback_query.answer()

    # Устанавливаем состояние ожидания файла
    user_waiting_for_file[callback_query.from_user.id] = "finance"

    await callback_query.message.edit_text(
        "📤 <b>Загрузка финансового отчета WB</b>\n\n"
        "📋 Отправьте Excel файл с финансовым отчетом, скачанный из личного кабинета Wildberries.\n\n"
        "💡 <b>Как получить файл:</b>\n"
        "1. Зайдите в личный кабинет WB\n"
        '2. Перейдите в "Финансы" → "Отчеты"\n'
        "3. Выберите нужный период\n"
        "4. Скачайте Excel файл\n"
        "5. Отправьте его сюда\n\n"
        "⏳ <b>Жду файл...</b>",
        parse_mode="HTML",
    )


# Себестоимость
@dp.callback_query_handler(lambda c: c.data == "cost_template")
async def cost_template_callback(callback_query: types.CallbackQuery):
    """Шаблон себестоимости"""
    await callback_query.answer()
    await cost_template_handler(callback_query.message)


@dp.callback_query_handler(lambda c: c.data == "cost_summary")
async def cost_summary_callback(callback_query: types.CallbackQuery):
    """Сводка себестоимости"""
    await callback_query.answer()
    await cost_summary_handler(callback_query.message)


# === ОБРАБОТЧИК ЗАГРУЗКИ EXCEL ФАЙЛОВ ===

# Хранилище для отслеживания ожидания файлов
user_waiting_for_file = {}  # {user_id: file_type}


@dp.message_handler(content_types=["document"])
async def handle_document_upload(message: types.Message):
    """Обработка загруженных Excel файлов"""
    user_id = message.from_user.id

    # Проверяем, ожидает ли пользователь загрузку файла
    if user_id not in user_waiting_for_file:
        await message.reply(
            "❓ Я не ожидал файл. Используйте меню 'Загрузка данных WB' для начала."
        )
        return

    file_type = user_waiting_for_file[user_id]
    document = message.document

    # Проверяем тип файла
    if not document.file_name.endswith((".xlsx", ".xls")):
        await message.reply("❌ Пожалуйста, загрузите файл Excel (.xlsx или .xls)")
        return

    try:
        # Отправляем сообщение о начале обработки
        processing_msg = await message.reply("⏳ Обрабатываю файл...")

        # Скачиваем файл
        file_info = await bot.get_file(document.file_id)
        file_path = f"uploads/wb_reports/{document.file_name}"

        # Создаем директорию если не существует
        import os

        os.makedirs("uploads/wb_reports", exist_ok=True)

        # Загружаем файл
        await bot.download_file(file_info.file_path, file_path)

        # Обрабатываем в зависимости от типа
        if file_type == "sales":
            analysis = await wb_excel_processor.process_sales_report(file_path)
        elif file_type == "orders":
            analysis = await wb_excel_processor.process_orders_report(file_path)
        elif file_type == "finance":
            analysis = await wb_excel_processor.process_finance_report(file_path)
        else:
            analysis = {"success": False, "error": "Неизвестный тип файла"}

        # Форматируем и отправляем результат
        report = wb_excel_processor.format_analysis_report(analysis)

        await processing_msg.edit_text(
            report + "\n\n💡 <i>Файл успешно обработан и сохранен для дальнейшего анализа</i>",
            parse_mode="HTML",
        )

        # Удаляем пользователя из ожидания
        del user_waiting_for_file[user_id]

        # Удаляем временный файл
        try:
            os.remove(file_path)
        except:
            pass

    except Exception as e:
        logger.error(f"Ошибка обработки файла: {e}")
        await message.reply(f"❌ Ошибка обработки файла: {e}")

        # Удаляем пользователя из ожидания
        if user_id in user_waiting_for_file:
            del user_waiting_for_file[user_id]


if __name__ == "__main__":
    # Запуск бота
    executor.start_polling(dp, on_startup=on_startup, on_shutdown=on_shutdown, skip_updates=True)
